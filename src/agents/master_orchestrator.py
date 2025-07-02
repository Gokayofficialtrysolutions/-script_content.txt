import asyncio, json, requests, subprocess, threading, queue, time, datetime
import torch
import aiohttp
from diffusers import DiffusionPipeline
from moviepy.editor import VideoFileClip
from pydub import AudioSegment
import pyttsx3
import shutil
from concurrent.futures import ThreadPoolExecutor,ProcessPoolExecutor
from dataclasses import dataclass, field
from typing import List,Dict,Any,Optional, Callable, Coroutine, Tuple # Added Tuple
from pathlib import Path
# Removed direct transformers import as NLU is now agent-based
# from transformers import pipeline as hf_pipeline
import sys
import re
import platform

try:
    from tools.auto_dev import auto_dev
except ImportError:
    try:
        print("Attempting standard import for tools.auto_dev")
        from tools.auto_dev import auto_dev
    except ImportError as e_imp:
        print(f"CRITICAL: Failed to import auto_dev from tools.auto_dev: {e_imp}")
        auto_dev = None

import chromadb
from chromadb.utils import embedding_functions
import uuid
from collections import defaultdict

import fitz  # PyMuPDF
import docx
import openpyxl
from bs4 import BeautifulSoup
import io # For BytesIO

from duckduckgo_search import DDGS

from collections import defaultdict
from ..core import prompt_constructors
from ..core.rl_logger import RLExperienceLogger
from ..core.async_tools import AsyncTask, AsyncTaskStatus
from ..core.event_system import SystemEvent
from ..core.conversation_history import ConversationTurn, ConversationContextManager
from ..core.kb_schemas import BaseKBSchema, PlanExecutionRecordDC, CodeExplanationDC, WebServiceScrapeResultDC, GenericDocumentDC # New KB Schema Imports
from ..core.knowledge_graph import KnowledgeGraph # Import KnowledgeGraph
from ..core.rl_policy_manager import RLPolicyManager # Import RLPolicyManager
from ..nlu_processing.nlu_engine import NLUProcessor, NLUProcessingError # Import NLUProcessor
from ..nlu_processing.nlu_config import nlu_default_config # Import default NLU config
from ..nlu_processing.nlu_results import NLUResult # Import NLUResult for type hinting


# --- New/Modified Dataclasses for Service Definitions ---
@dataclass
class AgentServiceParameter:
    name: str
    type: str  # e.g., "string", "integer", "float", "boolean", "list[string]", "dict"
    required: bool
    description: str
    default_value: Optional[Any] = None

@dataclass
class AgentServiceReturn:
    type: str # e.g., "string", "integer", "dict", "list[CustomObject]"
    description: str

@dataclass
class AgentServiceDefinition:
    name: str
    description: str
    parameters: List[AgentServiceParameter]
    returns: AgentServiceReturn
    handler_method_name: Optional[str] = None

@dataclass
class Agent:
   name:str
   model:str
   specialty:str
   active:bool=True
   estimated_complexity:Optional[str]=None
   estimated_speed:Optional[str]=None # Added estimated_speed
   # `provided_services` is read from JSON but not stored on the Agent object itself.
   # It's processed by the orchestrator into self.service_definitions.


class TerminusOrchestrator:
   def __init__(self):
       self.agents: List[Agent] = []
       self.service_definitions: Dict[Tuple[str, str], AgentServiceDefinition] = {}
       self.service_handlers: Dict[Tuple[str, str], Callable[..., Coroutine[Any, Any, Dict]]] = {}

       # --- Structures for Asynchronous Task Management ---
       # self.active_async_tasks: Stores the actual asyncio.Task objects for currently running tasks, keyed by task_id.
       # These are the raw tasks that are being awaited by the system.
       self.active_async_tasks: Dict[str, asyncio.Task] = {}
       # self.async_task_registry: Stores AsyncTask dataclass instances, providing a persistent record
       # of each task's state, result, or error, keyed by task_id. This registry holds the history
       # and final outcomes of tasks even after the asyncio.Task object is done.
       self.async_task_registry: Dict[str, AsyncTask] = {}
       # self._task_registry_lock: An asyncio.Lock to ensure thread-safe (or rather, async-safe)
       # modifications to async_task_registry and active_async_tasks.
       self._task_registry_lock = asyncio.Lock()
       # --- End Async Task Management Structures ---

       # --- Event Bus Structures ---
       # self.event_subscribers: Maps event_type strings to a list of asynchronous callback handlers.
       self.event_subscribers: Dict[str, List[Callable[[SystemEvent], Coroutine[Any, Any, None]]]] = defaultdict(list)
       # self.event_processing_queue: An asyncio.Queue where published SystemEvent objects are placed before dispatch.
       self.event_processing_queue: asyncio.Queue[SystemEvent] = asyncio.Queue()
       # self._event_dispatcher_task: Holds the asyncio.Task object for the _event_dispatcher_loop. Started in __init__.
       self._event_dispatcher_task: Optional[asyncio.Task] = None
       # --- End Event Bus Structures ---

       self.install_dir = Path(__file__).resolve().parent.parent
       self.agents_config_path = self.install_dir / "agents.json"
       self.models_config_path = self.install_dir / "models.conf"
       self.data_dir = self.install_dir / "data"
       self.logs_dir = self.install_dir / "logs"
       self.tools_dir = self.install_dir / "tools"
       self.generated_images_dir = self.data_dir / "generated_images"
       self.video_processing_dir = self.data_dir / "video_outputs"
       self.audio_processing_dir = self.data_dir / "audio_outputs"
       self.chroma_db_path = str(self.data_dir / "vector_store")
       self.feedback_log_file_path = self.logs_dir / "feedback_log.jsonl"
       self.feedback_analyzer_script_path = self.tools_dir / "feedback_analyzer.py"

       # Initialize Conversation Context Manager with configurable parameters
       # TODO: Load these from a dedicated orchestrator configuration file in the future.
       self.conv_ctx_mgr_config = {
           "summarization_threshold_turns": 20,
           "summarization_chunk_size": 10,
           "min_turns_to_keep_raw_at_end": 5,
           # For get_contextual_history defaults:
           "default_max_tokens": 3000,
           "default_desired_recent_turns": 7
       }
       self.conversation_context_manager = ConversationContextManager(
           summarization_threshold_turns=self.conv_ctx_mgr_config["summarization_threshold_turns"],
           summarization_chunk_size=self.conv_ctx_mgr_config["summarization_chunk_size"],
           min_turns_to_keep_raw_at_end=self.conv_ctx_mgr_config["min_turns_to_keep_raw_at_end"]
       )
       self.conversation_history: List[ConversationTurn] = [] # Now stores ConversationTurn objects

       for dir_path in [self.data_dir, self.logs_dir, self.tools_dir,
                        self.generated_images_dir, self.video_processing_dir,
                        self.audio_processing_dir, Path(self.chroma_db_path).parent]:
           dir_path.mkdir(parents=True, exist_ok=True)

       try:
           # Agent configuration loading
           with open(self.agents_config_path, 'r', encoding='utf-8') as f:
               agents_data_raw = json.load(f)

           for agent_config_raw in agents_data_raw:
               # Separate provided_services from other agent fields
               provided_services_raw = agent_config_raw.pop("provided_services", [])

               # Create Agent instance without provided_services
               agent_instance = Agent(**agent_config_raw)
               self.agents.append(agent_instance)

               # Process and store service definitions
               for service_def_raw in provided_services_raw:
                   try:
                       # Deserialize parameters
                       params_data = service_def_raw.get("parameters", [])
                       service_params = [AgentServiceParameter(**p) for p in params_data]

                       # Deserialize returns
                       returns_data = service_def_raw.get("returns", {})
                       service_returns = AgentServiceReturn(**returns_data)

                       service_def_obj = AgentServiceDefinition(
                           name=service_def_raw["name"],
                           description=service_def_raw["description"],
                           parameters=service_params,
                           returns=service_returns,
                           handler_method_name=service_def_raw.get("handler_method_name")
                       )

                       service_key = (agent_instance.name, service_def_obj.name)
                       self.service_definitions[service_key] = service_def_obj
                       print(f"Loaded service definition: {agent_instance.name} -> {service_def_obj.name}")

                       if service_def_obj.handler_method_name:
                           handler_method = getattr(self, service_def_obj.handler_method_name, None)
                           if handler_method and callable(handler_method):
                               self.service_handlers[service_key] = handler_method
                               print(f"  Registered direct handler '{service_def_obj.handler_method_name}' for service '{service_def_obj.name}'.")
                           else:
                               print(f"  WARNING: Handler method '{service_def_obj.handler_method_name}' for service '{service_def_obj.name}' not found or not callable on orchestrator.")
                   except Exception as e_service:
                       print(f"ERROR loading service definition for agent {agent_instance.name}: {service_def_raw.get('name', 'UNKNOWN_SERVICE')}. Details: {e_service}")

       except FileNotFoundError:
           print(f"CRITICAL ERROR: agents.json not found at {self.agents_config_path}. No agents or services loaded.")
       except json.JSONDecodeError as e_json:
           print(f"CRITICAL ERROR: Failed to decode agents.json: {e_json}. No agents or services loaded.")
       except Exception as e_outer:
           print(f"CRITICAL ERROR loading agents configuration: {e_outer}. Some agents/services may not be loaded.")

       self.ollama_url="http://localhost:11434/api/generate"
       self.image_gen_pipeline = None
       self.device = "cuda" if torch.cuda.is_available() else "cpu"
       self.image_gen_model_id = "stabilityai/stable-diffusion-xl-base-1.0"

       try:
           self.tts_engine = pyttsx3.init()
       except Exception as e:
           print(f"WARNING: Failed to initialize TTS engine: {e}.")
           self.tts_engine = None

       self.candidate_intent_labels = [
           "image_generation", "code_generation", "code_modification", "code_explanation",
           "project_scaffolding", "video_info", "video_frame_extraction", "video_to_gif",
           "audio_info", "audio_format_conversion", "text_to_speech",
           "data_analysis", "web_search", "document_processing", "general_question_answering",
           "complex_task_planning", "system_information_query", "knowledge_base_query",
           "feedback_submission", "feedback_analysis_request", "agent_service_call"
       ]

       self.conversation_history = []
       self.max_history_items = 10
       self.kb_collection_name = "terminus_knowledge_v1"
       self.knowledge_collection = None
       try:
           self.chroma_client = chromadb.PersistentClient(path=self.chroma_db_path)
           default_ef = embedding_functions.SentenceTransformerEmbeddingFunction()
           self.knowledge_collection = self.chroma_client.get_or_create_collection(name=self.kb_collection_name, embedding_function=default_ef)
           print(f"KB initialized: Collection '{self.kb_collection_name}' at {self.chroma_db_path}.")
       except Exception as e: print(f"CRITICAL ERROR initializing ChromaDB: {e}. KB unavailable.")

       # Old message bus attributes are removed as the new event system replaces them.
       # self.message_bus_subscribers = defaultdict(list)
       # self.message_processing_tasks = set()
       # self._setup_initial_event_listeners() # This call is removed. Subscriptions will use the new event bus.

       # Start the new event dispatcher loop
       self._event_dispatcher_task = asyncio.create_task(self._event_dispatcher_loop())
       print("TerminusOrchestrator: New event dispatcher loop started.")

       # Subscribe initial internal event handlers
       self.subscribe_to_event("kb.content.added", self._event_handler_kb_content_added)
       self.subscribe_to_event("user.feedback.submitted", self._event_handler_log_user_feedback) # Added for 2.4

       # self.service_handlers is now populated dynamically during agent loading
       # based on "handler_method_name" in service definitions.
       # Old static definition:
       # self.service_handlers = {
       # ("CodeMaster", "validate_code_syntax"): self._service_codemaster_validate_syntax
       # }
       self.default_high_priority_retries = 1
       print(f"TerminusOrchestrator initialized. {len(self.service_handlers)} direct service handlers registered. Default high-priority retries: {self.default_high_priority_retries}.")

       self.rl_experience_log_path = self.logs_dir / "rl_experience_log.jsonl"
       self.rl_logger = RLExperienceLogger(self.rl_experience_log_path)
       print(f"RL Experience Logger initialized. Logging to: {self.rl_experience_log_path}")

       # Initialize Knowledge Graph
       self.knowledge_graph_db_path = self.data_dir / "terminus_graph.db"
       try:
           self.kg_instance = KnowledgeGraph(db_path=self.knowledge_graph_db_path)
           print(f"KnowledgeGraph instance initialized with DB at: {self.knowledge_graph_db_path}")
       except Exception as e_kg:
           self.kg_instance = None
           print(f"CRITICAL ERROR initializing KnowledgeGraph: {e_kg}. Graph features will be unavailable.")

       # Initialize RL Policy Manager
       self.rl_policy_storage_path = self.data_dir / "rl_action_preferences.json"
       try:
           # Pass the orchestrator itself as the event_bus and the path to the log file
           self.rl_policy_manager = RLPolicyManager(
               policy_storage_path=self.rl_policy_storage_path,
               event_bus=self, # The orchestrator instance acts as the event bus provider
               experience_log_file_path=self.rl_experience_log_path
           )
           print(f"RLPolicyManager initialized. Policy storage: {self.rl_policy_storage_path}. Listening for log events.")
       except Exception as e_rl_init:
           self.rl_policy_manager = None
           print(f"CRITICAL ERROR initializing RLPolicyManager: {e_rl_init}. RL-based strategy selection and event-driven updates will be disabled.")

       # Initialize NLU Processor
       try:
           self.nlu_processor = NLUProcessor(config=nlu_default_config) # Using default config for now
           print("NLUProcessor initialized successfully.")
       except NLUProcessingError as e_nlu_init:
           self.nlu_processor = None
           print(f"CRITICAL ERROR initializing NLUProcessor: {e_nlu_init}. NLU features will be unavailable.")
           # Optionally, re-raise or handle more gracefully depending on how critical NLU is at startup
       except Exception as e_nlu_generic:
            self.nlu_processor = None
            print(f"CRITICAL UNEXPECTED ERROR initializing NLUProcessor: {e_nlu_generic}. NLU features will be unavailable.")

   async def _orchestrate_conversation_summarization(self):
        """
        Checks if conversation history needs summarization and orchestrates it.
        This method calls an LLM agent to perform the summarization if needed.
        It should be called periodically or before fetching context for long conversations.
        """
        if not hasattr(self, 'conversation_context_manager'): # Safety check
            print("[Orchestrator] ERROR: ConversationContextManager not initialized. Skipping summarization.")
            return

        chunk_data = self.conversation_context_manager.identify_chunk_for_summarization()
        if not chunk_data:
            # print("[Orchestrator] No conversation chunk identified for summarization at this time.")
            return

        original_indices, turns_to_summarize = chunk_data
        if not turns_to_summarize:
            return

        print(f"[Orchestrator] Attempting to summarize {len(turns_to_summarize)} turns (indices {original_indices}).")

        # Prepare text for summarization
        text_to_summarize = "\n".join([f"{t.role}: {t.content}" for t in turns_to_summarize])

        summarizer_agent = next((a for a in self.agents if a.name == "DocSummarizer" and a.active), None)
        if not summarizer_agent:
            summarizer_agent = next((a for a in self.agents if a.name == "GeneralPurposeAgent" and a.active), None) # Fallback

        if not summarizer_agent:
            print("[Orchestrator] ERROR: No summarizer agent (DocSummarizer or GeneralPurposeAgent) available. Cannot summarize conversation.")
            return

        summary_prompt = (
            f"Summarize the following conversation excerpt concisely, capturing key information, decisions, and outcomes. "
            f"The summary will replace these turns in a longer conversation history.\n\n"
            f"EXCERPT:\n{text_to_summarize[:10000]}\n\nSUMMARY:" # Limit input to LLM
        )

        # Call execute_agent for summarization. This will return a task_id.
        summary_task_submission = await self.execute_agent(summarizer_agent, summary_prompt)

        if summary_task_submission.get("status") == "pending_async":
            summary_task_id = summary_task_submission["task_id"]
            print(f"[Orchestrator] Conversation summarization task {summary_task_id} submitted. Awaiting result...")

            summary_text = None
            # Await the result of the summarization task
            while True:
                await asyncio.sleep(0.5) # Poll interval for summary task
                task_info = await self.get_async_task_info(summary_task_id)
                if not task_info:
                    print(f"[Orchestrator] ERROR: Summary task {summary_task_id} info not found.")
                    return # Cannot proceed with summarization

                if task_info.status == AsyncTaskStatus.COMPLETED:
                    if task_info.result and task_info.result.get("status") == "success":
                        summary_text = task_info.result.get("response")
                        print(f"[Orchestrator] Summary task {summary_task_id} completed. Summary: '{summary_text[:100]}...'")
                    else:
                        print(f"[Orchestrator] ERROR: Summary task {summary_task_id} completed but failed or no response: {task_info.result}")
                    break
                elif task_info.status == AsyncTaskStatus.FAILED:
                    print(f"[Orchestrator] ERROR: Summary task {summary_task_id} failed: {task_info.error}")
                    break
                elif task_info.status == AsyncTaskStatus.CANCELLED:
                    print(f"[Orchestrator] WARN: Summary task {summary_task_id} was cancelled.")
                    break

            if summary_text and summary_text.strip():
                # Determine start and end times for the summary description
                start_time_str = turns_to_summarize[0].timestamp.strftime('%Y-%m-%d %H:%M')
                end_time_str = turns_to_summarize[-1].timestamp.strftime('%Y-%m-%d %H:%M')

                summary_turn_content = f"Summary of conversation from {start_time_str} to {end_time_str} UTC: {summary_text}"
                summary_turn = ConversationTurn(
                    role="system",
                    content=summary_turn_content,
                    metadata={
                        "is_summary": True,
                        "summarized_turn_count": len(turns_to_summarize),
                        # Storing original indices might be complex if _managed_history shifts often.
                        # Storing first/last original turn timestamps is safer.
                        "summarized_period_start_ts": turns_to_summarize[0].timestamp.isoformat(),
                        "summarized_period_end_ts": turns_to_summarize[-1].timestamp.isoformat(),
                    }
                )
                self.conversation_context_manager.replace_turns_with_summary(original_indices, summary_turn)
                print(f"[Orchestrator] Conversation history updated with summary turn.")
            else:
                print(f"[Orchestrator] Failed to obtain a valid summary. No changes made to conversation history.")
        else:
            # This case (execute_agent not returning pending_async for an LLM agent) should be rare now.
            print(f"[Orchestrator] ERROR: Summarization call to agent {summarizer_agent.name} did not return pending_async status: {summary_task_submission}")


   # --- Asynchronous Task Management Methods ---
   async def _async_task_wrapper(self, task_id: str, coro: Coroutine, task_name: Optional[str]):
        """
        Internal wrapper for submitted coroutines. Handles the lifecycle of an AsyncTask:
        1. Sets status to RUNNING in async_task_registry.
        2. Awaits the coroutine.
        3. On completion, sets status to COMPLETED and stores the result.
        4. On failure (exception), sets status to FAILED and stores the error.
        5. On cancellation, sets status to CANCELLED.
        6. Removes the corresponding asyncio.Task from active_async_tasks.
        All updates to async_task_registry are performed under _task_registry_lock.
        """
        async with self._task_registry_lock:
            task_info = self.async_task_registry.get(task_id)
            if not task_info:
                # This case should ideally not be reached if submit_async_task correctly creates the entry.
                # However, as a safeguard:
                print(f"[AsyncTask-{task_id}] WARNING: Task info not found in registry at start of wrapper. Creating.")
                task_info = AsyncTask(task_id=task_id, name=task_name or "Unnamed Task")
                self.async_task_registry[task_id] = task_info

            task_info.status = AsyncTaskStatus.RUNNING
            task_info.started_at = datetime.datetime.utcnow() # Use UTC for consistency
            print(f"[AsyncTask-{task_id} ({task_info.name})] Status changed to RUNNING.")

        try:
            result = await coro
            async with self._task_registry_lock:
                task_info.status = AsyncTaskStatus.COMPLETED
                task_info.result = result
                task_info.completed_at = datetime.datetime.utcnow() # Use UTC
                print(f"[AsyncTask-{task_id} ({task_info.name})] Status changed to COMPLETED.")
        except asyncio.CancelledError:
            async with self._task_registry_lock:
                task_info.status = AsyncTaskStatus.CANCELLED
                task_info.error = "Task was cancelled."
                task_info.completed_at = datetime.datetime.utcnow() # Use UTC
                print(f"[AsyncTask-{task_id} ({task_info.name})] Status changed to CANCELLED.")
            # Propagate CancelledError if the calling context needs to handle it (e.g. asyncio.gather)
            # For now, we're consuming it here as the wrapper's role is to update status.
        except Exception as e:
            async with self._task_registry_lock:
                task_info.status = AsyncTaskStatus.FAILED
                task_info.error = f"{type(e).__name__}: {str(e)}"
                task_info.completed_at = datetime.datetime.utcnow() # Use UTC
                print(f"[AsyncTask-{task_id} ({task_info.name})] Status changed to FAILED. Error: {task_info.error}")
        finally:
            # Remove the asyncio.Task object from active_async_tasks as it's no longer running.
            # The AsyncTask entry in async_task_registry remains for history/result retrieval.
            self.active_async_tasks.pop(task_id, None)
            print(f"[AsyncTask-{task_id} ({task_info.name})] Wrapper finished, removed from active tasks list.")


   async def submit_async_task(self, coro: Coroutine, name: Optional[str] = None) -> str:
        """
        Submits a coroutine to be run as an asyncio.Task, managed by the orchestrator.
        Returns a unique task ID.
        """
        task_id = str(uuid.uuid4())

        async with self._task_registry_lock:
            task_info = AsyncTask(task_id=task_id, name=name or "Unnamed Task", status=AsyncTaskStatus.PENDING)
            self.async_task_registry[task_id] = task_info
            print(f"[AsyncTask-{task_id} ({task_info.name})] Status changed to PENDING and registered.")

        # Create and store the asyncio.Task, wrapped by our _async_task_wrapper
        # The wrapper will handle updating the status to RUNNING, COMPLETED, or FAILED.
        asyncio_task = asyncio.create_task(self._async_task_wrapper(task_id, coro, name))
        self.active_async_tasks[task_id] = asyncio_task

        # Store the asyncio.Task object in the AsyncTask dataclass for potential advanced control (e.g., cancellation)
        # This is done after ensuring task_info is in the registry.
        async with self._task_registry_lock:
             self.async_task_registry[task_id]._task_obj = asyncio_task # Assigning to private field

        print(f"[Orchestrator] Submitted AsyncTask ID: {task_id} for '{name}'. Coroutine: {coro.__name__ if hasattr(coro, '__name__') else type(coro).__name__}")
        return task_id

   async def get_async_task_info(self, task_id: str) -> Optional[AsyncTask]:
        """
        Retrieves the AsyncTask dataclass instance for a given task_id.
        This provides full info including status, result, error.
        """
        async with self._task_registry_lock: # Ensure consistent read
            task_info = self.async_task_registry.get(task_id)

        if task_info:
            return task_info
        else:
            print(f"[Orchestrator] No task info found for ID: {task_id}")
            return None

   async def cancel_async_task(self, task_id: str) -> bool: # For future use
        """
        Attempts to cancel an active asynchronous task.
        """
        async with self._task_registry_lock:
            task_info = self.async_task_registry.get(task_id)
            asyncio_task_obj = self.active_async_tasks.get(task_id)

            if not task_info or not asyncio_task_obj:
                print(f"[AsyncTask-{task_id}] Cannot cancel: Task not found or not active.")
                return False

            if task_info.status in [AsyncTaskStatus.COMPLETED, AsyncTaskStatus.FAILED, AsyncTaskStatus.CANCELLED]:
                print(f"[AsyncTask-{task_id}] Cannot cancel: Task already in terminal state ({task_info.status.name}).")
                return False

            cancelled = asyncio_task_obj.cancel()
            if cancelled:
                # The _async_task_wrapper will handle setting status to CANCELLED
                print(f"[AsyncTask-{task_id}] Cancellation requested.")
                # Optionally, could immediately update status to PENDING_CANCELLATION here
            else:
                print(f"[AsyncTask-{task_id}] Cancellation request failed (task may already be completing).")
            return cancelled
   # --- End Asynchronous Task Management Methods ---

   # --- Event Bus Methods ---
   async def publish_event(self, event_type: str, source_component: str, payload: Dict) -> str:
        """
        Creates a SystemEvent and puts it onto the event_processing_queue.
        Returns the event_id.
        """
        event = SystemEvent(
            event_type=event_type,
            source_component=source_component,
            payload=payload
        )
        await self.event_processing_queue.put(event)
        print(f"[EventBus] Published Event ID: {event.event_id}, Type: '{event.event_type}', Source: '{event.source_component}', Payload: {str(payload)[:100]}...")
        return event.event_id

   def subscribe_to_event(self, event_type: str, handler: Callable[[SystemEvent], Coroutine[Any, Any, None]]):
        """
        Subscribes an asynchronous handler to a specific event type.
        """
        self.event_subscribers[event_type].append(handler)
        handler_name = getattr(handler, '__name__', str(handler))
        print(f"[EventBus] Subscribed handler '{handler_name}' to event type '{event_type}'.")

   async def _event_dispatcher_loop(self):
        """
        Continuously gets events from the queue and dispatches them to subscribers.
        """
        print("[EventBus] Dispatcher loop started. Waiting for events...")
        while True:
            try:
                event = await self.event_processing_queue.get()
                print(f"[EventBus] Dispatching Event ID: {event.event_id}, Type: '{event.event_type}'...")

                handlers_to_call = self.event_subscribers.get(event.event_type, [])
                if not handlers_to_call:
                    print(f"[EventBus] No subscribers for event type '{event.event_type}'. Event ID: {event.event_id}")
                    self.event_processing_queue.task_done()
                    continue

                # Using asyncio.create_task for each handler to allow them to run concurrently
                # and not block the dispatcher loop if one handler is slow or errors.
                handler_tasks = []
                for handler in handlers_to_call:
                    handler_name = getattr(handler, '__name__', str(handler))
                    print(f"[EventBus] Calling handler '{handler_name}' for Event ID: {event.event_id} Type: {event.event_type}")
                    handler_tasks.append(asyncio.create_task(self._safe_execute_handler(handler, event)))

                # Wait for all handlers for this event to complete (or error)
                # This is optional; if we don't want to wait, we can just launch tasks.
                # For now, let's gather them to see their completion.
                await asyncio.gather(*handler_tasks, return_exceptions=False) # Errors handled in _safe_execute_handler

                self.event_processing_queue.task_done()
                print(f"[EventBus] Finished processing Event ID: {event.event_id}, Type: '{event.event_type}'.")

            except asyncio.CancelledError:
                print("[EventBus] Dispatcher loop cancelled.")
                break
            except Exception as e:
                print(f"[EventBus] CRITICAL ERROR in dispatcher loop: {e}. Loop continues but this event might be lost.")
                # Potentially re-queue the event or log to a dead-letter queue in a real system
                # For now, just mark as done to avoid blocking queue if get() was successful.
                if 'event' in locals() and self.event_processing_queue._unfinished_tasks > 0 : # Check if get() was successful
                     self.event_processing_queue.task_done()


   async def _safe_execute_handler(self, handler: Callable[[SystemEvent], Coroutine[Any, Any, None]], event: SystemEvent):
        """
        Safely executes a single event handler, catching and logging exceptions.
        """
        handler_name = getattr(handler, '__name__', str(handler))
        try:
            await handler(event)
            print(f"[EventBusHandler] Handler '{handler_name}' completed successfully for Event ID: {event.event_id}.")
        except Exception as e:
            print(f"[EventBusHandler] ERROR: Handler '{handler_name}' failed for Event ID: {event.event_id}. Error: {type(e).__name__}: {e}")
            # Add more detailed logging, e.g., traceback.format_exc() if needed
   # --- End Event Bus Methods ---


   def get_agent_capabilities_description(self) -> str:
       descriptions = []
       for a in self.agents:
           if a.active:
               complexity_info = f" (Complexity: {a.estimated_complexity})" if a.estimated_complexity else ""
               speed_info = f" (Speed: {a.estimated_speed})" if a.estimated_speed else ""
               descriptions.append(f"- {a.name}: Specializes in '{a.specialty}'. Uses model: {a.model}.{complexity_info}{speed_info}")
       return "\n".join(descriptions) if descriptions else "No active agents available."

    async def _handle_system_event(self, event: SystemEvent): # Signature changed to SystemEvent
       # This method will be re-subscribed to relevant events in Phase 2 if still needed.
       # For now, its direct subscriptions via _setup_initial_event_listeners are removed.
       print(f"[LegacyEventHandler] Event ID: {event.event_id}, Type: '{event.event_type}', Src: '{event.source_component}', Payload: {event.payload}")

   async def _event_handler_log_user_feedback(self, event: SystemEvent):
        """
        Simple event handler to log when user feedback is submitted.
        """
        feedback_id = event.payload.get("feedback_id", "N/A")
        item_id = event.payload.get("item_id", "N/A")
        rating = event.payload.get("rating", "N/A")
        print(f"[UserFeedbackLogger] Received 'user.feedback.submitted' event. Feedback ID: {feedback_id}, Item ID: {item_id}, Rating: {rating}. Event ID: {event.event_id}")

   # _setup_initial_event_listeners REMOVED
   # publish_message REMOVED
   # subscribe_to_message REMOVED

   async def _event_handler_kb_content_added(self, event: SystemEvent):
       """
       Handles the 'kb.content.added' event to trigger content analysis.
       Formerly _handle_new_kb_content_for_analysis.
       """
       kb_id = event.payload.get("kb_id")
       # Extract metadata and schema_type from the event payload, which should have the richer metadata.
       event_metadata = event.payload.get("metadata", {})
       schema_type = event_metadata.get("kb_schema_type")
       source_op = event_metadata.get("source", event.payload.get("source_operation", "unknown_source"))

       handler_id = f"[ContentAnalysisHandler event_id:{event.event_id} kb_id:{kb_id} schema:{schema_type} src_op:{source_op}]"

       # Skip analysis for certain source operations or schema types that don't need generic keyword/topic extraction.
       if source_op in ["feedback_analysis_report"] or schema_type in ["FeedbackReport"]: # Example, adjust as needed
           print(f"{handler_id} INFO: Skipping content analysis for KB item from source '{source_op}' or schema '{schema_type}'.")
           return

       print(f"{handler_id} START: Processing event type: {event.event_type}")
       if self.knowledge_collection is None:
           print(f"{handler_id} ERROR: Knowledge base not available. Skipping analysis."); return
       if not kb_id:
           print(f"{handler_id} ERROR: No valid kb_id in event payload. Cannot process."); return

       try:
           # Fetch the item from ChromaDB to get its document string and existing full metadata
           item_data_from_db = self.knowledge_collection.get(ids=[kb_id], include=["documents", "metadatas"])
           if not (item_data_from_db and item_data_from_db.get('ids') and item_data_from_db['ids'][0]):
               print(f"{handler_id} ERROR: KB item '{kb_id}' not found in ChromaDB for analysis."); return

           doc_json_string = item_data_from_db['documents'][0] if item_data_from_db.get('documents') and item_data_from_db['documents'][0] else None
           db_metadata = item_data_from_db['metadatas'][0] if item_data_from_db.get('metadatas') and item_data_from_db['metadatas'][0] else {}

           # Re-confirm schema_type from DB metadata as source of truth for what's stored.
           # Event payload schema_type is what was intended at storage time.
           db_schema_type = db_metadata.get('kb_schema_type', schema_type) # Prefer DB if available, else from event

           if not doc_json_string:
               print(f"{handler_id} INFO: KB item '{kb_id}' has empty document string in DB. Skipping analysis."); return

           # Check if already analyzed by this mechanism (using DB metadata)
           if db_metadata.get("analysis_by_agent") == "ContentAnalysisAgent" and \
              (db_metadata.get("extracted_keywords") or db_metadata.get("extracted_topics")):
               print(f"{handler_id} INFO: Content for KB ID '{kb_id}' (schema: {db_schema_type}) already analyzed by ContentAnalysisAgent. Skipping re-analysis.")
               return

           text_for_analysis = ""
           if db_schema_type == "PlanExecutionRecord":
               try:
                   plan_rec = PlanExecutionRecordDC.from_json_string(doc_json_string)
                   text_for_analysis = f"User Request: {plan_rec.original_user_request}\nOutcome Summary: {plan_rec.final_summary_to_user}"
                   print(f"{handler_id} INFO: Extracted text from PlanExecutionRecord for analysis.")
               except Exception as e_plan_parse:
                   print(f"{handler_id} WARNING: Could not parse PlanExecutionRecord for KB ID '{kb_id}'. Using raw JSON string. Error: {e_plan_parse}")
                   text_for_analysis = doc_json_string # Fallback to raw JSON string
           elif db_schema_type == "WebServiceScrapeResult":
               try:
                   scrape_res = WebServiceScrapeResultDC.from_json_string(doc_json_string)
                   text_for_analysis = f"Title: {scrape_res.title}\nSummary: {scrape_res.main_content_summary}"
                   print(f"{handler_id} INFO: Extracted text from WebServiceScrapeResult for analysis.")
               except Exception as e_scrape_parse:
                   print(f"{handler_id} WARNING: Could not parse WebServiceScrapeResult for KB ID '{kb_id}'. Using raw JSON string. Error: {e_scrape_parse}")
                   text_for_analysis = doc_json_string # Fallback
           elif db_schema_type == "GenericDocument":
               try:
                   gen_doc = GenericDocumentDC.from_json_string(doc_json_string)
                   # Primarily analyze the summary, but could also include a snippet of original if desired.
                   text_for_analysis = f"Source: {gen_doc.source_identifier}\nSummary: {gen_doc.summary_content}"
                   if gen_doc.original_content and len(gen_doc.original_content) < 2000: # Only add short original content
                       text_for_analysis += f"\nOriginal Content Snippet: {gen_doc.original_content[:200]}..."
                   print(f"{handler_id} INFO: Extracted text from GenericDocument (summary and optional original snippet) for analysis.")
               except Exception as e_gendoc_parse:
                   print(f"{handler_id} WARNING: Could not parse GenericDocument for KB ID '{kb_id}'. Using raw JSON string. Error: {e_gendoc_parse}")
                   text_for_analysis = doc_json_string # Fallback
           elif db_schema_type == "CodeExplanation": # Code explanations might be better served by their own keywords
                print(f"{handler_id} INFO: Skipping generic keyword/topic analysis for CodeExplanation schema '{db_schema_type}'. Assumed to have its own keywords.")
                return
           else: # Default to using the full document string (which might be plain text or JSON of unknown schema)
               text_for_analysis = doc_json_string
               if db_schema_type:
                   print(f"{handler_id} INFO: Unknown schema '{db_schema_type}'. Analyzing raw document content.")
               else:
                   print(f"{handler_id} INFO: No schema type. Analyzing raw document content.")


           if not text_for_analysis.strip():
               print(f"{handler_id} INFO: No text content derived for analysis from KB ID '{kb_id}' (schema: {db_schema_type}). Skipping."); return

           analysis_agent = next((a for a in self.agents if a.name == "ContentAnalysisAgent" and a.active), None)
           if not analysis_agent:
               print(f"{handler_id} ERROR: ContentAnalysisAgent not found/active. Skipping."); return

           analysis_prompt = (f"Analyze the following text content derived from a knowledge base item (ID: {kb_id}, Schema: {db_schema_type or 'N/A'}):\n---\n{text_for_analysis[:15000]}\n---\n"
                              f"Provide output as a JSON object with 'keywords' (comma-separated string of 2-5 relevant keywords, or 'NONE') "
                              f"and 'topics' (1-3 comma-separated strings describing main topics, or 'NONE').\n"
                              f"Example: {{\"keywords\": \"k1, k2, k3\", \"topics\": \"Topic A, Topic B\"}}\nJSON Output:")

           print(f"{handler_id} INFO: Calling LLM for content analysis of KB ID '{kb_id}'.")
           llm_call_or_task = await self.execute_agent(analysis_agent, analysis_prompt)

           llm_result: Optional[Dict] = None
           if llm_call_or_task.get("status") == "pending_async":
                analysis_task_id = llm_call_or_task["task_id"]
                print(f"{handler_id} LLM analysis for KB '{kb_id}' submitted as task {analysis_task_id}. Awaiting result...")
                while True: # This handler will block for its own LLM call.
                    await asyncio.sleep(0.2)
                    task_info = await self.get_async_task_info(analysis_task_id)
                    if not task_info:
                        print(f"{handler_id} ERROR: LLM analysis task {analysis_task_id} info not found for KB '{kb_id}'."); return
                    if task_info.status == AsyncTaskStatus.COMPLETED:
                        llm_result = task_info.result
                        break
                    elif task_info.status == AsyncTaskStatus.FAILED:
                        print(f"{handler_id} ERROR: LLM analysis task {analysis_task_id} for KB '{kb_id}' failed: {task_info.error}"); return
                    elif task_info.status == AsyncTaskStatus.CANCELLED:
                         print(f"{handler_id} ERROR: LLM analysis task {analysis_task_id} for KB '{kb_id}' cancelled."); return
           else:
                llm_result = llm_call_or_task

           if not llm_result or not (llm_result.get("status") == "success" and llm_result.get("response","").strip()):
               print(f"{handler_id} ERROR: LLM analysis call for KB '{kb_id}' failed or produced no response. Status: {llm_result.get('status') if llm_result else 'N/A'}, Resp: {llm_result.get('response') if llm_result else 'N/A'}"); return

           print(f"{handler_id} SUCCESS: LLM analysis for KB '{kb_id}' successful.")
           llm_response_str = llm_result.get("response").strip()
           extracted_keywords, extracted_topics = "", ""
           try:
               data = json.loads(llm_response_str)
               raw_kw = data.get("keywords","").strip()
               extracted_keywords = raw_kw if raw_kw and raw_kw.upper() != "NONE" else ""
               raw_tp = data.get("topics","").strip()
               extracted_topics = raw_tp if raw_tp and raw_tp.upper() != "NONE" else ""
           except json.JSONDecodeError:
               print(f"{handler_id} WARNING: Failed to parse LLM JSON for content analysis of KB '{kb_id}'. Raw: '{llm_response_str}'. Using raw as keywords if applicable and not 'none'.")
               if "keywords" not in llm_response_str.lower() and "topics" not in llm_response_str.lower() and llm_response_str.upper() != "NONE":
                   extracted_keywords = llm_response_str[:200]

           if extracted_keywords or extracted_topics:
               new_meta = {
                   "analysis_by_agent": analysis_agent.name,
                   "analysis_model_used": analysis_agent.model,
                   "analysis_timestamp_iso": datetime.datetime.utcnow().isoformat()
                }
               if extracted_keywords: new_meta["extracted_keywords"] = extracted_keywords
               if extracted_topics: new_meta["extracted_topics"] = extracted_topics

               print(f"{handler_id} INFO: Attempting metadata update for KB '{kb_id}' with keywords: '{extracted_keywords}', topics: '{extracted_topics}'.")
               update_status = await self._update_kb_item_metadata(kb_id, new_meta)
               if update_status.get("status") == "success":
                   print(f"{handler_id} SUCCESS: Metadata update for KB '{kb_id}' successful.")
                   # Also update Knowledge Graph with keywords and topics
                   if self.kg_instance:
                       # Ensure KB item node exists
                       self.kg_instance.add_node(node_id=kb_id, node_type=db_schema_type or "GenericContent", content_preview=text_for_analysis[:100])

                       if extracted_keywords:
                           for kw in extracted_keywords.split(','):
                               kw_clean = kw.strip().lower()
                               if kw_clean:
                                   kw_node_id = f"keyword_{kw_clean.replace(' ', '_')}"
                                   self.kg_instance.add_node(node_id=kw_node_id, node_type="Keyword", content_preview=kw_clean)
                                   self.kg_instance.add_edge(source_node_id=kb_id, target_node_id=kw_node_id, relationship_type="HAS_KEYWORD", ensure_nodes=False)
                           print(f"{handler_id} INFO: Added/Updated keyword relationships in KG for KB ID '{kb_id}'.")

                       if extracted_topics:
                           for topic in extracted_topics.split(','):
                               topic_clean = topic.strip().lower()
                               if topic_clean:
                                   topic_node_id = f"topic_{topic_clean.replace(' ', '_')}"
                                   self.kg_instance.add_node(node_id=topic_node_id, node_type="Topic", content_preview=topic_clean)
                                   self.kg_instance.add_edge(source_node_id=kb_id, target_node_id=topic_node_id, relationship_type="HAS_TOPIC", ensure_nodes=False)
                           print(f"{handler_id} INFO: Added/Updated topic relationships in KG for KB ID '{kb_id}'.")
               else:
                   print(f"{handler_id} ERROR: Metadata update for KB '{kb_id}' failed. Msg: {update_status.get('message')}")
           else:
               print(f"{handler_id} INFO: No keywords or topics extracted for KB '{kb_id}'. No metadata update, no KG keyword/topic edges added.")
       except Exception as e:
           print(f"{handler_id} UNHANDLED ERROR in content analysis handler for KB '{kb_id}': {type(e).__name__}: {e}")
       finally:
           print(f"{handler_id} END: Finished processing event for KB '{kb_id}'.")

   async def publish_message(self, message_type: str, source_agent_name: str, payload: Dict) -> str: #This is part of the old message bus, will be removed.
       message_id = str(uuid.uuid4()); message = { "message_id": message_id, "message_type": message_type, "source_agent_name": source_agent_name, "timestamp_iso": datetime.datetime.now().isoformat(), "payload": payload }
       print(f"[MessageBus] Publishing: ID={message_id}, Type='{message_type}', Src='{source_agent_name}'")
       for handler in list(self.message_bus_subscribers.get(message_type, [])):
            try:
                if asyncio.iscoroutinefunction(handler): asyncio.create_task(handler(message)).add_done_callback(self.message_processing_tasks.discard)
                elif isinstance(handler, asyncio.Queue): await handler.put(message)
            except Exception as e: print(f"ERROR dispatching message {message_id} to {handler}: {e}")
       return message_id

   def subscribe_to_message(self, message_type: str, handler: Callable[..., Coroutine[Any, Any, None]] | asyncio.Queue):
       self.message_bus_subscribers[message_type].append(handler)
       print(f"[MessageBus] Subscribed '{getattr(handler, '__name__', str(type(handler)))}' to '{message_type}'.")

   async def _execute_single_plan_step(self, step_definition: Dict, full_plan_list: List[Dict], current_step_outputs: Dict) -> Dict:
       step_id = step_definition.get("step_id"); agent_name = step_definition.get("agent_name")
       task_prompt = step_definition.get("task_prompt", ""); dependencies = step_definition.get("dependencies", [])
       output_var_name = step_definition.get("output_variable_name")
       step_priority = step_definition.get("priority", "normal").lower()
       explicit_max_retries = step_definition.get("max_retries")
       if explicit_max_retries is not None: max_retries = explicit_max_retries
       elif step_priority == "high": max_retries = self.default_high_priority_retries
       else: max_retries = 0
       retry_delay_seconds = step_definition.get("retry_delay_seconds", 5)
       retry_on_statuses = step_definition.get("retry_on_statuses", ["error"])
       current_execution_retries = 0
       log_prefix = f"[Priority: {step_priority.upper()}] " if step_priority == "high" else ""
       target_agent = next((a for a in self.agents if a.name == agent_name and a.active), None)
       if not target_agent: return {"status": "error", "agent": agent_name, "step_id": step_id, "response": f"Agent '{agent_name}' not found/active."}
       while True:
           current_task_prompt = task_prompt
           for dep_id in dependencies:
               dep_key = next((p.get("output_variable_name",f"step_{dep_id}_output") for p in full_plan_list if p.get("step_id")==dep_id),None)
               if dep_key and dep_key in current_step_outputs:
                   val = current_step_outputs[dep_key]
                   if isinstance(val,dict):
                       for m in re.finditer(r"{{{{("+re.escape(dep_key)+r")\.(\w+)}}}}",current_task_prompt):
                           if m.group(2) in val: current_task_prompt=current_task_prompt.replace(m.group(0),str(val[m.group(2)]))
                   current_task_prompt=current_task_prompt.replace(f"{{{{{dep_key}}}}}",str(val))
           retry_info = f" (Retry {current_execution_retries}/{max_retries})" if max_retries > 0 and current_execution_retries > 0 else ""
           if max_retries > 0 and current_execution_retries == 0 and not retry_info: retry_info = f" (Attempt 1/{max_retries + 1})"
           print(f"{log_prefix}Executing step {step_id}{retry_info}: Agent='{agent_name}', Prompt='{current_task_prompt[:100]}...'")
           step_result = await self.execute_agent(target_agent, current_task_prompt)
           if step_result.get("status") == "success":
               key_to_store = output_var_name if output_var_name else f"step_{step_id}_output"
               current_step_outputs[key_to_store] = step_result.get("response")
               for mk in ["image_path","frame_path","gif_path","speech_path","modified_file"]:
                   if mk in step_result: current_step_outputs[f"{key_to_store}_{mk}"]=step_result[mk]
               print(f"{log_prefix}Step {step_id} completed successfully.")
               return step_result
           print(f"{log_prefix}Step {step_id} failed on attempt {current_execution_retries + 1}/{max_retries + 1}. Status: {step_result.get('status')}, Response: {str(step_result.get('response'))[:100]}...")
           current_execution_retries += 1
           if not (current_execution_retries <= max_retries and step_result.get("status") in retry_on_statuses):
               print(f"{log_prefix}Step {step_id} exhausted retries or status not retryable. Final status: {step_result.get('status')}")
               return step_result
           print(f"{log_prefix}Step {step_id} retrying in {retry_delay_seconds}s...")
           await asyncio.sleep(retry_delay_seconds)

   async def _handle_agent_service_call(self, service_call_step_def: Dict, current_step_outputs: Dict, full_plan_list: List[Dict]) -> Dict:
        """
        Handles a plan step of type 'agent_service_call'.
        It resolves parameters, validates them against the service definition,
        dispatches to a direct Python handler if available, or falls back to an LLM call
        to the target agent. Manages synchronous and asynchronous results from these calls.
        """
        step_id = service_call_step_def.get("step_id", "unknown_service_call_step")
        step_priority = service_call_step_def.get("priority", "normal").lower()
        log_prefix = f"[{self.__class__.__name__}._handle_agent_service_call Step-{step_id}]" # More specific log prefix
        if step_priority == "high": log_prefix += f" [Priority: HIGH]"
        target_agent_name = service_call_step_def.get("target_agent_name")
        service_name = service_call_step_def.get("service_name")
        service_params_template = service_call_step_def.get("service_params", {})
        output_var_name = service_call_step_def.get("output_variable_name")

        if not all([target_agent_name, service_name]):
            return {"step_id": step_id, "status": "error", "response": "Missing target_agent_name or service_name for agent_service_call.", "data": None, "error_code": "MISSING_SERVICE_INFO"}

        service_key = (target_agent_name, service_name)
        service_def = self.service_definitions.get(service_key)

        if not service_def:
            # Fallback for services not formally defined (legacy or dynamic)
            # This part remains similar to original, but consider if this fallback should be stricter or logged more prominently.
            print(f"{log_prefix}WARNING: Service definition for '{service_name}' on agent '{target_agent_name}' not found. Proceeding with LLM fallback without formal validation.")
            # Proceed with LLM fallback as before, but without parameter validation based on a definition.
            # For brevity, the old parameter resolution and LLM call logic is assumed here if service_def is None.
            # This section would essentially replicate the old logic path for LLM fallback.
            # However, for this refactoring, we will prioritize the new structured approach.
            # If a service is called, it SHOULD have a definition.
            # For now, let's make it an error if not defined, to enforce the new structure.
            error_message = f"Service '{service_name}' on agent '{target_agent_name}' is not defined."
            print(f"{log_prefix} ERROR: {error_message}")
            return {"step_id": step_id, "status": "error", "response": error_message, "data": None, "error_code": "SERVICE_NOT_DEFINED"}

        # --- Parameter Resolution, Validation, and Coercion ---
        # This section iterates through parameters defined in the service_def,
        # resolves their values from the plan step's service_params (handling templates),
        # applies defaults, checks for required parameters, and performs basic type coercion.
        resolved_params = {}
        param_validation_errors = []

        for defined_param in service_def.parameters:
            param_name = defined_param.name
            raw_value = service_params_template.get(param_name) # Get parameter value from plan step

            # 1. Resolve from template string if applicable (e.g., "{{step_X_output.some_key}}")
            if isinstance(raw_value, str):
                substituted_value = raw_value
                # Regex to find placeholders like {{variable.path}} or {{variable}}
                for dep_match in re.finditer(r"{{{{([\w.-]+)}}}}", raw_value):
                    var_path = dep_match.group(1)
                    val_to_sub = current_step_outputs.get(var_path) # Direct lookup first

                    # Handle dot-notation for nested dictionary access from previous step outputs
                    if '.' in var_path and val_to_sub is None:
                        base_key, *attrs = var_path.split('.')
                        if base_key in current_step_outputs:
                            temp_val = current_step_outputs[base_key]
                            try:
                                for part in attrs:
                                    if isinstance(temp_val, dict): temp_val = temp_val.get(part)
                                    elif isinstance(temp_val, list) and part.isdigit(): temp_val = temp_val[int(part)]
                                    else: temp_val = None; break
                                if temp_val is not None: val_to_sub = temp_val
                            except Exception as e_attr: # pylint: disable=broad-except
                                print(f"{log_prefix} Warning: Error accessing attribute '{part}' in '{var_path}': {e_attr}")
                                val_to_sub = None # Ensure it's None if path is invalid

                    if val_to_sub is not None:
                        substituted_value = substituted_value.replace(dep_match.group(0), str(val_to_sub))
                    else:
                        print(f"{log_prefix} Warning: Could not resolve template variable '{var_path}' for parameter '{param_name}'. It might be optional or resolved later if the template itself is the value.")
                actual_value = substituted_value
            else: # Value is literal (e.g. int, bool, list, dict from JSON plan) or already resolved by prior step.
                actual_value = raw_value

            # 2. Handle missing parameters: use default if available, else error if required.
            if actual_value is None: # Parameter not provided in plan step's service_params
                if defined_param.default_value is not None:
                    actual_value = defined_param.default_value
                    print(f"{log_prefix} INFO: Using default value for parameter '{param_name}': {actual_value}")
                elif defined_param.required:
                    param_validation_errors.append(f"Required parameter '{param_name}' is missing and has no default.")
                    continue # Skip to next parameter for validation
                else: # Optional parameter without a value and no default; skip it.
                    continue

            # 3. Type Coercion (Basic types for now)
            coerced_value = actual_value
            target_type = defined_param.type.lower()
            try:
                if actual_value is not None: # Only coerce if there's a value
                    if target_type == "integer": coerced_value = int(actual_value)
                    elif target_type == "float": coerced_value = float(actual_value)
                    elif target_type == "boolean":
                        if isinstance(actual_value, str):
                            coerced_value = actual_value.lower() in ["true", "1", "yes"]
                        else:
                            coerced_value = bool(actual_value)
                    elif target_type == "string": # Ensure it's a string if specified
                        coerced_value = str(actual_value)
                    # For "list", "dict", we assume the value from plan/template is already in the correct complex type.
                    # More sophisticated validation for list/dict item types could be added later if needed.
            except ValueError as e_coerce:
                param_validation_errors.append(f"Parameter '{param_name}' (value: '{actual_value}') could not be coerced to type '{defined_param.type}': {e_coerce}")
                continue # Skip to next parameter for validation

            resolved_params[param_name] = coerced_value

        if param_validation_errors:
            error_msg = f"Parameter validation failed for service '{service_name}' on agent '{target_agent_name}': " + "; ".join(param_validation_errors)
            print(f"{log_prefix} ERROR: {error_msg}")
            return {"step_id": step_id, "status": "error", "response": error_msg, "data": None, "error_code": "PARAMETER_VALIDATION_FAILED"}

        # --- Execute Service (Direct Handler or LLM Fallback) ---
        service_result_structured: Dict[str, Any] # This will hold the standardized service output

        if service_key in self.service_handlers: # Check for a direct Python handler
            handler_method = self.service_handlers[service_key]
            print(f"{log_prefix} Executing direct handler '{handler_method.__name__}' for service '{service_name}' with params: {resolved_params}")
            try:
                # Direct handlers are expected to be async and accept 'params' and 'service_definition'.
                # They should return the standardized dictionary: {"status": ..., "data": ..., "message": ..., "error_code": ...}
                service_result_structured = await handler_method(params=resolved_params, service_definition=service_def)

                # If the direct handler itself returns a 'pending_async' status (e.g., if it internally started another managed task),
                # propagate this status. This allows handlers to be fully non-blocking if they choose.
                if service_result_structured.get("status") == "pending_async" and "task_id" in service_result_structured:
                    print(f"{log_prefix} Direct handler for service '{service_name}' returned a pending task: {service_result_structured['task_id']}")
                    return { # Ensure the propagated structure is consistent for plan execution
                        "step_id": step_id,
                        "agent_name": f"{target_agent_name} (Service: {service_name})", # Or from service_result_structured if provided
                        "status": "pending_async",
                        "task_id": service_result_structured["task_id"],
                        "response": service_result_structured.get("message", f"Service '{service_name}' on '{target_agent_name}' (via direct handler) initiated as async task."),
                        "data": None
                    }
            except Exception as e_handler: # Catch errors from the handler execution itself
                err_msg = f"Direct handler '{handler_method.__name__}' for service '{service_name}' raised an exception: {type(e_handler).__name__}: {e_handler}"
                print(f"{log_prefix} ERROR: {err_msg}")
                service_result_structured = {"status": "error", "data": None, "message": err_msg, "error_code": "HANDLER_EXCEPTION"}
        else: # No direct handler, proceed to LLM fallback
            print(f"{log_prefix} No direct handler for service '{service_name}' on agent '{target_agent_name}'. Using LLM fallback.")
            target_agent = next((a for a in self.agents if a.name == target_agent_name and a.active), None)
            if not target_agent:
                err_msg = f"Target agent '{target_agent_name}' for LLM fallback of service '{service_name}' not found or inactive."
                print(f"{log_prefix} ERROR: {err_msg}")
                return {"step_id": step_id, "status": "error", "response": err_msg, "data": None, "error_code": "AGENT_NOT_FOUND"}

            # Construct the detailed prompt for the LLM agent based on the service definition
            param_details_for_prompt = "\n".join([f"  - {p.name} ({p.type}, {'required' if p.required else 'optional'}{', default: '+str(p.default_value) if p.default_value is not None and not p.required else ''}): {p.description}" for p in service_def.parameters])
            returns_details_for_prompt = f"  - type: {service_def.returns.type}\n  - description: {service_def.returns.description}"

            fallback_prompt = (
                f"You are agent '{target_agent_name}'. You need to perform the service called '{service_def.name}'.\n"
                f"Service Description: {service_def.description}\n\n"
                f"Parameters Expected by the service:\n{param_details_for_prompt}\n\n"
                f"Expected Return Structure from you (the LLM Agent):\n"
                f"  A JSON object containing:\n"
                f"  - 'status': (string) 'success' or 'error'.\n"
                f"  - 'data': (matching '{service_def.returns.type}') The actual data payload as described: {service_def.returns.description}.\n"
                f"  - 'message': (string, optional) A human-readable message about the operation.\n\n"
                f"Parameters Provided for this specific call:\n{json.dumps(resolved_params, indent=2)}\n\n"
                f"Process the provided parameters according to the service description and generate a JSON response that strictly conforms to the 'Expected Return Structure' detailed above."
            )

            # Call execute_agent, which might return a 'pending_async' status if the target is an LLM agent
            llm_call_result_or_task = await self.execute_agent(target_agent, fallback_prompt)

            if llm_call_result_or_task.get("status") == "pending_async":
                print(f"{log_prefix} LLM fallback for service '{service_name}' initiated as task: {llm_call_result_or_task['task_id']}")
                return { # Propagate the pending_async status
                    "step_id": step_id,
                    "agent_name": f"{target_agent_name} (Service: {service_name}, LLM Fallback)",
                    "status": "pending_async",
                    "task_id": llm_call_result_or_task["task_id"],
                    "response": llm_call_result_or_task.get("message", f"LLM fallback for service '{service_name}' initiated."),
                    "data": None
                }
            # If execute_agent returned a synchronous result (e.g., an immediate error before task submission)
            elif llm_call_result_or_task.get("status") == "success":
                # This case is less expected now for LLM agents, as execute_agent should return pending_async.
                # However, if it occurs, try to parse the response as the structured JSON.
                print(f"{log_prefix} WARNING: LLM fallback for service '{service_name}' returned synchronous success. Expecting pending_async for LLM agents.")
                try:
                    parsed_llm_response = json.loads(llm_call_result_or_task.get("response","{}"))
                    if not (isinstance(parsed_llm_response, dict) and "status" in parsed_llm_response and "data" in parsed_llm_response):
                        raise ValueError("LLM fallback (sync success) response does not conform to expected structured JSON with status and data keys.")
                    service_result_structured = parsed_llm_response
                    service_result_structured.setdefault("message", f"LLM for service '{service_name}' (sync fallback) completed with status: {service_result_structured['status']}")
                except Exception as e_parse_llm_sync:
                    err_msg = f"Error parsing synchronous LLM fallback response for service '{service_name}': {e_parse_llm_sync}. Raw: {llm_call_result_or_task.get('response','')[:200]}..."
                    print(f"{log_prefix} ERROR: {err_msg}")
                    service_result_structured = {"status": "error", "data": None, "message": err_msg, "error_code": "LLM_FALLBACK_SYNC_RESPONSE_PARSE_ERROR"}
            else: # Synchronous error from execute_agent itself
                err_msg = f"LLM fallback call (via execute_agent) for service '{service_name}' failed before async submission: {llm_call_result_or_task.get('response')}"
                print(f"{log_prefix} ERROR: {err_msg}")
                service_result_structured = {"status": "error", "data": None, "message": err_msg, "error_code": "LLM_FALLBACK_PRE_ASYNC_CALL_FAILED"}

        # --- Process final result (if the step was not pending_async) ---
        final_status = service_result_structured.get("status", "error")
        final_data = service_result_structured.get("data")
        final_message = service_result_structured.get("message", "Service call completed.")
        error_code = service_result_structured.get("error_code")

        if final_status != "success":
            print(f"{log_prefix}Service call step {step_id} ('{service_name}' on '{target_agent_name}') failed. Status: {final_status}, Message: {final_message[:100]}...")
        else:
            print(f"{log_prefix}Service call step {step_id} ('{service_name}' on '{target_agent_name}') completed successfully.")

        if output_var_name and final_status == "success":
            current_step_outputs[output_var_name] = final_data

        return {
            "step_id": step_id,
            "agent_name": f"{target_agent_name} (Service: {service_name})",
            "status": final_status,
            "response": final_message, # User-facing/log summary message
            "data": final_data, # Actual data payload
            "error_code": error_code
        }

   async def _service_codemaster_validate_syntax(self, params: Dict, service_definition: AgentServiceDefinition) -> Dict:
        # service_definition is now passed but not strictly needed if params are already validated by caller.
        # It could be used for more complex logic if the handler serves multiple similar services.
        code_snippet = params.get("code_snippet")
        # language = params.get("language", "python") # Default already handled by param validation if defined so.
        # The 'language' parameter would have been resolved by _handle_agent_service_call using defaults from service_def if not provided.
        language = params.get("language")


        codemaster_agent = next((a for a in self.agents if a.name == "CodeMaster" and a.active), None)
        if not codemaster_agent:
            return {"status": "error", "data": None, "message": "CodeMaster agent not available for syntax validation.", "error_code": "AGENT_UNAVAILABLE"}

        prompt = (f"Analyze the following {language} code snippet for syntax errors. Respond ONLY with a JSON object containing two keys: 'is_valid' (boolean) and 'errors' (a list of strings, empty if valid). Do not add any explanatory text outside the JSON.\nCode:\n```\n{code_snippet}\n```\nJSON Response:")

        llm_response_or_task = await self.execute_agent(codemaster_agent, prompt)

        llm_response: Dict
        if llm_response_or_task.get("status") == "pending_async":
            task_id = llm_response_or_task["task_id"]
            print(f"[_service_codemaster_validate_syntax] LLM task {task_id} submitted. Awaiting result...")
            while True:
                await asyncio.sleep(0.1) # Quick poll
                task_info = await self.get_async_task_info(task_id)
                if not task_info:
                     return {"status": "error", "data": None, "message": f"Syntax validation LLM task {task_id} info not found.", "error_code": "ASYNC_TASK_NOT_FOUND"}

                if task_info.status == AsyncTaskStatus.COMPLETED:
                    if not isinstance(task_info.result, dict) or "status" not in task_info.result:
                        return {"status": "error", "data": None, "message": f"Syntax validation LLM task {task_id} completed with unexpected result format: {type(task_info.result)}.", "error_code": "LLM_RESULT_FORMAT_ERROR"}
                    llm_response = task_info.result
                    print(f"[_service_codemaster_validate_syntax] LLM task {task_id} completed.")
                    break
                elif task_info.status == AsyncTaskStatus.FAILED:
                    return {"status": "error", "data": None, "message": f"Syntax validation LLM task {task_id} failed: {task_info.error}", "error_code": "LLM_TASK_FAILED"}
                elif task_info.status == AsyncTaskStatus.CANCELLED:
                     return {"status": "error", "data": None, "message": f"Syntax validation LLM task {task_id} cancelled.", "error_code": "LLM_TASK_CANCELLED"}
        else:
            llm_response = llm_response_or_task

        if llm_response.get("status") == "success":
            try:
                validation_data = json.loads(llm_response.get("response"))
                if not isinstance(validation_data, dict) or "is_valid" not in validation_data or "errors" not in validation_data:
                     raise ValueError("LLM response for syntax validation is not a dict with 'is_valid' and 'errors' keys.")
                return {"status": "success", "data": validation_data, "message": f"Syntax validation for {language} completed. Valid: {validation_data.get('is_valid')}."}
            except json.JSONDecodeError:
                return {"status": "error", "data": None, "message": "CodeMaster (validator LLM) returned non-JSON response.", "raw_response": llm_response.get("response"), "error_code": "LLM_RESPONSE_MALFORMED"}
            except ValueError as e:
                 return {"status": "error", "data": None, "message": f"CodeMaster (validator LLM) response structure error: {e}", "raw_response": llm_response.get("response"), "error_code": "LLM_RESPONSE_STRUCTURE_INVALID"}
        else:
            return {"status": "error", "data": None, "message": f"CodeMaster LLM call failed for syntax validation: {llm_response.get('response')}", "error_code": "LLM_CALL_FAILED"}

   async def _service_docsummarizer_summarize_text(self, params: Dict, service_definition: AgentServiceDefinition) -> Dict:
        """
        Handles a service call to the DocSummarizer agent to summarize a given text.

        This method is invoked when a plan step specifies an `agent_service_call`
        targeting the 'DocSummarizer' agent and its 'summarize_text' service.
        It constructs a prompt for the DocSummarizer's underlying LLM, executes the
        agent, and handles the asynchronous response to return a structured
        service result containing the summary.

        Args:
            params (Dict): A dictionary containing the parameters for the service call.
                           Expected to have 'text_to_summarize'.
            service_definition (AgentServiceDefinition): The definition of the
                                                       'summarize_text' service.

        Returns:
            Dict: A structured dictionary indicating the outcome of the service call.
                  If successful and asynchronous, status will be 'pending_async' with a 'task_id'.
                  If completed synchronously (less likely for LLM), status 'success' with 'data'
                  containing the summary.
                  On error, status 'error' with an error message.
        """
        log_prefix = f"[{self.__class__.__name__}._service_docsummarizer_summarize_text]"
        text_to_summarize = params.get("text_to_summarize")

        if not text_to_summarize or not isinstance(text_to_summarize, str) or not text_to_summarize.strip():
            return {"status": "error", "data": None, "message": "Missing or empty 'text_to_summarize' parameter.", "error_code": "MISSING_PARAMETER"}

        summarizer_agent = next((a for a in self.agents if a.name == "DocSummarizer" and a.active), None)
        if not summarizer_agent:
            return {"status": "error", "data": None, "message": "DocSummarizer agent not available.", "error_code": "AGENT_UNAVAILABLE"}

        # Construct prompt for the DocSummarizer LLM
        # Limit input to avoid exceeding LLM context limits for very long texts.
        # The summarization task itself should be robust to this.
        max_input_length = 15000 # Example limit, could be configurable
        summary_prompt = (
            f"Please provide a concise summary of the following text. Focus on the key points and main information.\n\n"
            f"TEXT TO SUMMARIZE:\n```\n{text_to_summarize[:max_input_length]}\n```\n\nCONCISE SUMMARY:"
        )

        print(f"{log_prefix} Calling DocSummarizer agent for summarization. Text length: {len(text_to_summarize)} (truncated to {max_input_length} for prompt).")
        llm_response_or_task = await self.execute_agent(summarizer_agent, summary_prompt)

        # Handle the response from execute_agent (which might be pending_async)
        if llm_response_or_task.get("status") == "pending_async":
            task_id = llm_response_or_task["task_id"]
            print(f"{log_prefix} Summarization task {task_id} submitted. Propagating pending_async.")
            # Propagate the pending_async status as per the service call handling design
            return {
                "status": "pending_async",
                "task_id": task_id,
                "message": f"Summarization task initiated by DocSummarizer (task_id: {task_id})."
            }

        # If execute_agent returned a synchronous result (e.g., immediate error or direct success for non-LLM)
        # This path is less likely for an LLM-based summarizer but included for completeness.
        if llm_response_or_task.get("status") == "success":
            summary_text = llm_response_or_task.get("response")
            if summary_text and summary_text.strip():
                print(f"{log_prefix} Summarization (sync) successful.")
                # The 'data' field for this service should match the 'returns' type in agents.json, which is "string".
                # So, the summary_text itself is the data.
                return {"status": "success", "data": summary_text, "message": "Text summarized successfully (synchronous path)."}
            else:
                print(f"{log_prefix} Summarization (sync) returned empty response.")
                return {"status": "error", "data": None, "message": "Summarization (sync) resulted in an empty summary.", "error_code": "EMPTY_SUMMARY"}
        else: # Synchronous error from execute_agent
            error_msg = llm_response_or_task.get("response", "Unknown error during summarization agent call.")
            print(f"{log_prefix} Summarization agent call (sync) failed: {error_msg}")
            return {"status": "error", "data": None, "message": error_msg, "error_code": "LLM_CALL_FAILED_SYNC"}


   async def store_knowledge(
        self,
        content: Optional[str] = None, # Becomes optional if structured_content is primary
        metadata: Optional[Dict] = None,
        content_id: Optional[str] = None,
        schema_type: Optional[str] = None,      # New: e.g., "PlanExecutionRecord", "CodeExplanation"
        structured_content: Optional[BaseKBSchema] = None # New: Instance of a KB dataclass
    ) -> Dict:
        if self.knowledge_collection is None:
            return {"status": "error", "message": "KB not initialized."}

        final_id = content_id or str(uuid.uuid4())
        final_metadata = metadata.copy() if metadata else {} # Ensure we have a mutable dict
        document_content_for_chroma: str

        if structured_content:
            if not isinstance(structured_content, BaseKBSchema):
                return {"status": "error", "message": "structured_content is not a valid BaseKBSchema instance."}
            if not schema_type: # Infer schema_type if not explicitly provided but structured_content is
                schema_type = structured_content.__class__.__name__.replace("DC","") # e.g. PlanExecutionRecordDC -> PlanExecutionRecord
                print(f"[store_knowledge] INFO: Inferred schema_type as '{schema_type}' from structured_content.")

            document_content_for_chroma = structured_content.to_json_string()
            final_metadata['kb_schema_type'] = schema_type # Crucial for typed retrieval
            # Optionally, add other key fields from structured_content to metadata for direct filtering if needed
            # e.g., if structured_content is PlanExecutionRecordDC:
            # if schema_type == "PlanExecutionRecord" and hasattr(structured_content, 'status'):
            #     final_metadata['plan_status'] = structured_content.status
            # if schema_type == "PlanExecutionRecord" and hasattr(structured_content, 'primary_intent'):
            #     final_metadata['plan_intent'] = structured_content.primary_intent

        elif content is not None: # Standard text content
            document_content_for_chroma = content
            if schema_type: # If schema_type is provided for plain text, store it.
                 final_metadata['kb_schema_type'] = schema_type
        else:
            return {"status": "error", "message": "Either 'content' or 'structured_content' must be provided."}

        try:
            # Ensure metadata for ChromaDB contains only basic types
            chroma_meta_safe = {k: (str(v) if not isinstance(v, (str, int, float, bool)) else v) for k, v in final_metadata.items()}

            self.knowledge_collection.add(
                ids=[final_id],
                documents=[document_content_for_chroma],
                metadatas=[chroma_meta_safe] if chroma_meta_safe else [None]
            )

            # Publish event after successful storage
            event_payload = {
                "kb_id": final_id,
                "content_preview": document_content_for_chroma[:200] + "..." if len(document_content_for_chroma) > 200 else document_content_for_chroma,
                "metadata": final_metadata, # Send original, potentially richer metadata in event
                "source_operation": final_metadata.get("source", "unknown"),
                "schema_type": schema_type # Include schema_type in event
            }
            await self.publish_event(
                event_type="kb.content.added",
                source_component="TerminusOrchestrator.KnowledgeBase",
                payload=event_payload
            )
            print(f"[store_knowledge] Successfully stored KB ID: {final_id} (Schema: {schema_type or 'text'}) and published event.")
            return {"status": "success", "id": final_id, "message": "Content stored and event published."}
        except Exception as e:
            print(f"[store_knowledge] Error storing to KB or publishing event for ID {final_id}: {e}")
            return {"status": "error", "message": str(e)}

   async def retrieve_knowledge(self, query_text: str, n_results: int = 5, filter_metadata: Optional[Dict] = None) -> Dict:
        if self.knowledge_collection is None:
            return {"status": "error", "results": [], "message": "KB not initialized."}

        # Ensure filter_metadata uses values suitable for ChromaDB's 'where' clause
        # For this iteration, assuming direct string/numeric values in filter_metadata are fine.
        # Complex filters like $in, $nin might need specific construction if passed via filter_metadata.
        chroma_where_filter = filter_metadata if filter_metadata else None

        try:
            q_res = self.knowledge_collection.query(
                query_texts=[query_text] if query_text else None, # Allow query by filter only if query_text is None/empty
                n_results=max(1, n_results),
                where=chroma_where_filter,
                include=["documents", "metadatas", "distances"] # Ensure all are included
            )

            results = []
            if q_res and q_res.get('ids') and q_res['ids'][0]:
                for i, item_id in enumerate(q_res['ids'][0]):
                    doc_content_str = q_res['documents'][0][i] if q_res['documents'] and q_res['documents'][0] else None
                    metadata = q_res['metadatas'][0][i] if q_res['metadatas'] and q_res['metadatas'][0] else {}
                    distance = q_res['distances'][0][i] if q_res['distances'] and q_res['distances'][0] else None

                    result_item: Dict[str, Any] = {
                        "id": item_id,
                        "document_text": doc_content_str, # Always include the raw document string
                        "structured_document": None,      # Placeholder for deserialized object
                        "metadata": metadata,
                        "distance": distance
                    }

                    schema_type = metadata.get('kb_schema_type')
                    if schema_type and doc_content_str:
                        # Attempt to deserialize if schema_type is known
                        # This requires mapping schema_type string to the actual dataclass
                        schema_class_map = {
                            "PlanExecutionRecord": PlanExecutionRecordDC,
                            "CodeExplanation": CodeExplanationDC,
                            "WebServiceScrapeResult": WebServiceScrapeResultDC,
                            "GenericDocument": GenericDocumentDC,
                            # Add other schema types here as they are defined
                        }
                        TargetClass = schema_class_map.get(schema_type)
                        if TargetClass:
                            try:
                                result_item["structured_document"] = TargetClass.from_json_string(doc_content_str)
                                print(f"[retrieve_knowledge] Successfully deserialized KB ID {item_id} to schema {schema_type}.")
                            except Exception as e_deserialize:
                                print(f"[retrieve_knowledge] WARNING: Failed to deserialize KB ID {item_id} (schema: {schema_type}) from JSON string. Error: {e_deserialize}. Falling back to text.")
                        else:
                            print(f"[retrieve_knowledge] WARNING: Unknown kb_schema_type '{schema_type}' for KB ID {item_id}. Treating as text.")

                    results.append(result_item)

            return {"status": "success", "results": results}
        except Exception as e:
            print(f"[retrieve_knowledge] Error querying ChromaDB: {e}")
            return {"status": "error", "results": [], "message": str(e)}

   def store_user_feedback(self, item_id: str, item_type: str, rating: str, comment: Optional[str]=None, current_mode: Optional[str]=None, user_prompt_preview: Optional[str]=None) -> bool:
       try:
           feedback_id = str(uuid.uuid4())
           timestamp_iso = datetime.datetime.utcnow().isoformat() # Use UTC

           data_to_log = {
               "feedback_id": feedback_id,
               "timestamp_iso": timestamp_iso,
               "item_id": str(item_id),
               "item_type": str(item_type),
               "rating": str(rating),
               "comment": comment or "",
               "user_context":{
                   "operation_mode":current_mode,
                   "related_user_prompt_preview":user_prompt_preview[:200] if user_prompt_preview else None
                }
            }
           with open(self.feedback_log_file_path, 'a', encoding='utf-8') as f:
               f.write(json.dumps(data_to_log) + '\n')

           # Publish to the new event bus
           event_payload = {
               "feedback_id": feedback_id,
               "item_id": str(item_id),
               "item_type": str(item_type),
               "rating": str(rating),
               "comment_preview": (comment[:75] + "...") if comment and len(comment) > 75 else (comment or ""),
               "timestamp_iso": timestamp_iso
           }
           asyncio.create_task(self.publish_event(
               event_type="user.feedback.submitted",
               source_component="TerminusOrchestrator.FeedbackSystem", # Or simply "UserFeedbackSystem"
               payload=event_payload
           ))
           print(f"[store_user_feedback] Feedback {feedback_id} stored and event published.")
           return True
       except Exception as e:
           print(f"ERROR storing feedback or publishing event: {e}")
           return False

   async def generate_and_store_feedback_report(self) -> Dict:
       report_handler_id = "[FeedbackReport]"
       print(f"{report_handler_id} START: Generating and storing feedback analysis report.")
       if self.knowledge_collection is None: return {"status": "error", "message": "KB not initialized."}
       if not self.feedback_analyzer_script_path.exists(): return {"status": "error", "message": f"Analyzer script missing: {self.feedback_analyzer_script_path}"}
       try:
           print(f"{report_handler_id} INFO: Executing {self.feedback_analyzer_script_path} with log {self.feedback_log_file_path}")
           proc = await asyncio.create_subprocess_exec(sys.executable, str(self.feedback_analyzer_script_path), f"--log_file={self.feedback_log_file_path}", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
           stdout, stderr = await proc.communicate()
           if proc.returncode != 0:
               err_msg = stderr.decode().strip() if stderr else "Unknown error from feedback_analyzer.py"
               print(f"{report_handler_id} ERROR: Analyzer script failed. Code: {proc.returncode}. Err: {err_msg}"); return {"status": "error", "message": f"Analyzer script failed: {err_msg}"}
           print(f"{report_handler_id} SUCCESS: Analyzer script executed.")
           report_json_str = stdout.decode().strip()
           if not report_json_str: print(f"{report_handler_id} WARNING: Analyzer script produced no output."); return {"status": "error", "message":"Analyzer script empty output."}
           try: report_data = json.loads(report_json_str)
           except json.JSONDecodeError as e: print(f"{report_handler_id} ERROR: Failed to parse JSON from analyzer: {e}. Output: {report_json_str[:200]}..."); return {"status":"error", "message":f"Bad JSON from analyzer: {e}"}
           print(f"{report_handler_id} SUCCESS: Parsed report JSON.")
           expected_keys = ["report_id", "report_generation_timestamp_iso", "analysis_period_start_iso", "analysis_period_end_iso", "total_feedback_entries_processed", "overall_sentiment_distribution", "sentiment_by_item_type", "comment_previews"]
           missing_keys = [k for k in expected_keys if k not in report_data]
           if missing_keys: print(f"{report_handler_id} WARNING: Report JSON missing keys: {missing_keys}.")
           if "overall_sentiment_distribution" in report_data and not isinstance(report_data["overall_sentiment_distribution"],dict): print(f"{report_handler_id} WARNING: 'overall_sentiment_distribution' not a dict.")
           if "total_feedback_entries_processed" in report_data and not isinstance(report_data["total_feedback_entries_processed"],int): print(f"{report_handler_id} WARNING: 'total_feedback_entries_processed' not an int.")
           kb_meta = {"source":"feedback_analysis_report", "report_id":report_data.get("report_id",str(uuid.uuid4())), "report_date_iso":report_data.get("report_generation_timestamp_iso",datetime.datetime.now().isoformat()).split('T')[0], "analysis_start_iso":report_data.get("analysis_period_start_iso"), "analysis_end_iso":report_data.get("analysis_period_end_iso")}
           kb_meta = {k:v for k,v in kb_meta.items() if v is not None}
           print(f"{report_handler_id} INFO: Storing report in KB. Report ID: {kb_meta.get('report_id')}")
           store_res = await self.store_knowledge(content=report_json_str, metadata=kb_meta)
           if store_res.get("status") == "success":
               msg = f"Feedback report stored. KB ID: {store_res.get('id')}"
               print(f"{report_handler_id} SUCCESS: {msg}")
                # Event "kb.content.added" is already published by store_knowledge.
                # If a more specific "kb.feedback_report.added" event is desired, it can be published here too.
                # For now, we rely on the generic one. The metadata['source'] == 'feedback_analysis_report'
                # in the generic event's payload can be used by subscribers to differentiate.
                # Example of specific event:
                # await self.publish_event(
                # event_type="kb.feedback_report.added",
                # source_component="FeedbackAnalyzerSystem",
                #     payload={"report_id": kb_meta.get('report_id'), "kb_id": store_res.get("id")}
                # )
               # Add FeedbackReport node to Knowledge Graph and link if possible
               stored_kb_report_id = store_res.get("id")
               if self.kg_instance and stored_kb_report_id:
                   self.kg_instance.add_node(
                       node_id=stored_kb_report_id,
                       node_type="FeedbackReport",
                       content_preview=f"Report ID: {kb_meta.get('report_id')}, Overall Sentiment: {report_data.get('overall_sentiment_distribution',{}).get('positive',0)*100:.0f}% positive",
                       metadata={"report_id": kb_meta.get('report_id'), "generation_timestamp_iso": report_data.get("report_generation_timestamp_iso")}
                   )
                   print(f"{report_handler_id} SUCCESS: Added/Updated FeedbackReport node '{stored_kb_report_id}' in Knowledge Graph.")

                   # Conceptual: Attempt to link to a PlanExecutionRecord if item_id from feedback implies it.
                   # This requires feedback_log.jsonl to contain item_ids that are plan_log_kb_ids
                   # and feedback_analyzer.py to potentially aggregate feedback per item_id and include this
                   # 'source_item_id_for_feedback' in its report_data.
                   source_item_id_for_feedback = report_data.get("source_item_id_for_feedback_aggregation")
                   if source_item_id_for_feedback:
                       # Ensure the source item (e.g., a PlanLog) node exists or is created as a placeholder
                       # self.kg_instance.add_node(source_item_id_for_feedback, "UnknownSourceItem", content_preview="Source of feedback", ensure_nodes=True) # This would be ensure_nodes=True on add_edge

                       # Edge from FeedbackReport to the item it analyzes
                       self.kg_instance.add_edge(stored_kb_report_id, source_item_id_for_feedback,
                                                 "ANALYZES_FEEDBACK_FOR", ensure_nodes=True)
                       # Edge from the item to the FeedbackReport
                       self.kg_instance.add_edge(source_item_id_for_feedback, stored_kb_report_id,
                                                 "HAS_FEEDBACK_ANALYZED_IN", ensure_nodes=True)
                       print(f"{report_handler_id} INFO: Linked FeedbackReport '{stored_kb_report_id}' to source item '{source_item_id_for_feedback}' in KG.")

               return {"status":"success", "message":msg, "kb_id":stored_kb_report_id}
            else:
                print(f"{report_handler_id} ERROR: Failed to store report in KB. Msg: {store_res.get('message')}")
                return store_res
        except Exception as e:
            print(f"{report_handler_id} UNHANDLED ERROR: {e}")
            return {"status":"error","message":str(e)}
        finally:
            print(f"{report_handler_id} END: Processing finished.")

   async def _update_kb_item_metadata(self, kb_id: str, new_metadata_fields: Dict) -> Dict:
       if self.knowledge_collection is None: return {"status": "error", "message": "KB not initialized."}
       try:
           existing = self.knowledge_collection.get(ids=[kb_id], include=["metadatas","documents"])
           if not (existing and existing.get('ids') and existing['ids'][0]): return {"status":"error", "message":f"KB ID '{kb_id}' not found."}
           current_meta = existing['metadatas'][0] or {}; doc = existing['documents'][0]
           if doc is None: return {"status":"error", "message":f"Doc for KB ID '{kb_id}' missing."}
           updated_meta = current_meta.copy()
           for k,v in new_metadata_fields.items(): updated_meta[k] = str(v) if not isinstance(v,(str,int,float,bool)) else v
           self.knowledge_collection.update(ids=[kb_id],metadatas=[updated_meta],documents=[doc])
           return {"status":"success", "id":kb_id, "message":"Metadata updated."}
       except Exception as e: return {"status":"error", "message":str(e)}

   def get_conversation_history_for_display(self) -> List[Dict]:
       return list(self.conversation_history)

   async def _get_formatted_general_kb_context(self, kb_search_query: str, plan_handler_id: str) -> str:
       kb_general_ctx_str = ""
       if kb_search_query:
           print(f"[{plan_handler_id}] INFO: Querying KB for general context with: '{kb_search_query}'")
           general_hits = await self.retrieve_knowledge(
               kb_search_query,
               n_results=3,
               filter_metadata={"source": {"$nin": ["plan_execution_log", "feedback_analysis_report"]}}
           )
           if general_hits.get("status")=="success" and general_hits.get("results"):
               formatted_entries = [prompt_constructors.format_kb_entry_for_prompt(hit) for hit in general_hits["results"]]
               kb_general_ctx_str = "General Context from Knowledge Base (limit 3 relevance based on query):\n" + "\n".join(formatted_entries) + "\n\n"
               print(f"[{plan_handler_id}] INFO: Found {len(general_hits['results'])} general KB items.")
           else:
               print(f"[{plan_handler_id}] INFO: No general KB items found for query '{kb_search_query}'.")
       return kb_general_ctx_str

   async def _get_formatted_plan_log_insights(self, nlu_output: Dict, plan_handler_id: str) -> str:
       kb_plan_log_ctx_str = ""
       nlu_intent = nlu_output.get('intent','N/A')
       nlu_entities_str = str(nlu_output.get('entities',[]))[:50]
       plan_log_query = f"Intent: {nlu_intent}, Entities: {nlu_entities_str}"
       print(f"[{plan_handler_id}] INFO: Querying KB for plan logs with: '{plan_log_query}'")
       plan_log_hits = await self.retrieve_knowledge(plan_log_query, n_results=2, filter_metadata={"source":"plan_execution_log"})
       if plan_log_hits.get("status")=="success" and plan_log_hits.get("results"):
           formatted_plan_logs = [prompt_constructors.format_plan_log_entry_for_prompt(hit) for hit in plan_log_hits["results"]]
           kb_plan_log_ctx_str = "Insights from Past Plan Executions (limit 2 relevance based on NLU):\n" + "\n".join(formatted_plan_logs) + "\n\n"
           print(f"[{plan_handler_id}] INFO: Found {len(plan_log_hits['results'])} plan log items.")
       else:
           print(f"[{plan_handler_id}] INFO: No relevant plan logs found for query '{plan_log_query}'.")
       return kb_plan_log_ctx_str

   async def _get_formatted_feedback_insights(self, plan_handler_id: str) -> str:
       kb_feedback_ctx_str = ""
       feedback_query = "latest user feedback summary report"
       print(f"[{plan_handler_id}] INFO: Querying KB for feedback reports with: '{feedback_query}'")
       feedback_hits = await self.retrieve_knowledge(feedback_query, n_results=1, filter_metadata={"source":"feedback_analysis_report"})
       if feedback_hits.get("status")=="success" and feedback_hits.get("results"):
           formatted_feedback_reports = [prompt_constructors.format_feedback_report_for_prompt(hit) for hit in feedback_hits["results"]]
           kb_feedback_ctx_str = "Insights from Feedback Analysis Reports (limit 1 latest):\n" + "\n".join(formatted_feedback_reports) + "\n\n"
           print(f"[{plan_handler_id}] INFO: Found {len(feedback_hits['results'])} feedback report items.")
       else:
           print(f"[{plan_handler_id}] INFO: No feedback reports found for query '{feedback_query}'.")
       return kb_feedback_ctx_str

   def _construct_rl_state(self, user_prompt: str, nlu_output: Dict, kb_context_summary: Dict, previous_cycle_outcome: Optional[str]) -> Dict[str, Any]:
       state = {
           "nlu_intent": nlu_output.get("intent", "unknown"),
           "nlu_entities_count": len(nlu_output.get("entities", [])),
           "user_prompt_length_category": "short" if len(user_prompt) < 50 else "medium" if len(user_prompt) < 200 else "long",
           "kb_general_hits": kb_context_summary.get("general_hits_count", 0),
           "kb_plan_log_hits": kb_context_summary.get("plan_log_hits_count", 0),
           "kb_feedback_hits": kb_context_summary.get("feedback_hits_count", 0),
           "previous_cycle_outcome": previous_cycle_outcome if previous_cycle_outcome else "none",
       }
       return state

   def _calculate_rl_reward(self, execution_status: str, user_feedback_rating: Optional[str], num_revisions: int) -> float:
       reward = 0.0
       if execution_status == "success": reward += 0.5
       elif execution_status == "failure": reward -= 0.5
       if user_feedback_rating == "positive": reward += 1.0
       elif user_feedback_rating == "negative": reward -= 1.0
       reward -= num_revisions * 0.25
       return round(reward, 4)

   async def execute_master_plan(self, user_prompt: str, request_priority: Optional[str] = "normal") -> List[Dict]:
        """
        Orchestrates the execution of a multi-step plan generated by the MasterPlanner agent
        in response to a user prompt. Handles synchronous and asynchronous step execution,
        dependencies, conditional logic, loops (basic), and plan revisions.

        Args:
            user_prompt: The user's request.
            request_priority: The priority of the request (e.g., "normal", "high").

        Returns:
            A list of dictionaries, where each dictionary represents the final result of a plan step.
        """
       plan_handler_id = f"[MasterPlanner user_prompt:'{user_prompt[:50]}...' Priority:'{request_priority}']"
       rl_interaction_id = str(uuid.uuid4())
       timestamp_interaction_start = datetime.datetime.utcnow().isoformat() # Use UTC

       print(f"{plan_handler_id} START: Received request. RL Interaction ID: {rl_interaction_id}")

       # --- Perform NLU analysis on the user prompt ---
       nlu_result_for_plan: Optional[NLUResult] = None
       if self.nlu_processor:
           try:
               nlu_result_for_plan = self.nlu_processor.process_text(user_prompt) # This is a synchronous call
               print(f"{plan_handler_id} INFO: NLU processing complete. Intent: {nlu_result_for_plan.detected_intent.name if nlu_result_for_plan.detected_intent else 'N/A'}, Entities: {len(nlu_result_for_plan.entities)}")
               # Store NLU result in conversation history metadata for the user turn
               user_turn_metadata = {"nlu_result": nlu_result_for_plan.dict()} # Store as dict
           except Exception as e_nlu_proc:
               print(f"{plan_handler_id} WARNING: NLU processing failed during planning: {e_nlu_proc}")
               user_turn_metadata = {"nlu_processing_error": str(e_nlu_proc)}
       else:
           print(f"{plan_handler_id} WARNING: NLUProcessor not available. Proceeding without NLU analysis for this request.")
           user_turn_metadata = {"nlu_status": "unavailable"}

       # Add current user prompt to history as a ConversationTurn object
       # Extract keywords for the new user turn.
       user_prompt_keywords = self.conversation_context_manager.extract_keywords_from_text(user_prompt)
       user_turn = ConversationTurn(role="user", content=user_prompt, keywords=user_prompt_keywords, metadata=user_turn_metadata)
       self.conversation_history.append(user_turn)
       # print(f"{plan_handler_id} INFO: Extracted keywords from user prompt: {user_prompt_keywords}") # Already logged by NLU if available

       # Prune overall history if it exceeds max_history_items (simple FIFO for overall limit)
       if len(self.conversation_history) > self.max_history_items:
           self.conversation_history = self.conversation_history[-self.max_history_items:]

       # Update the context manager with the latest full history
       self.conversation_context_manager.update_full_history(self.conversation_history)

       # Orchestrate conversation summarization if needed
       await self._orchestrate_conversation_summarization()
       # After potential summarization, the context manager's internal _managed_history is updated.
       # get_contextual_history will now use this potentially summarized _managed_history.


       max_rev_attempts = 1; current_attempt = 0; plan_list = []; original_plan_json_str = ""
       final_exec_results = []
       step_outputs = {}
       first_attempt_nlu_output = {} # This will store the NLU result for the first planning attempt
       detailed_failure_ctx_for_rev = {}
       current_plan_log_kb_id = None
       state_for_executed_plan_log: Optional[Dict] = None
       action_for_executed_plan_log: Optional[str] = None
       prompt_details_for_executed_plan_log: Optional[Dict] = None

       # Use the NLU result obtained earlier if available
       if nlu_result_for_plan:
            # Convert NLUResult object to a dictionary format similar to what classify_user_intent (legacy) produced
            # This ensures `first_attempt_nlu_output` has a consistent structure for downstream use.
            first_attempt_nlu_output = {
                "status": "success", # Assuming success if we have an NLUResult object
                "intent": nlu_result_for_plan.detected_intent.name if nlu_result_for_plan.detected_intent else "unknown_intent",
                "intent_score": nlu_result_for_plan.detected_intent.confidence if nlu_result_for_plan.detected_intent else 0.0,
                "alternative_intents": [
                    {"intent": alt_name, "score": alt_score} # Use tuple directly
                    for alt_name, alt_score in (nlu_result_for_plan.detected_intent.alternate_intents or [])
                ] if nlu_result_for_plan.detected_intent else [],
                "entities": [ent.dict() for ent in nlu_result_for_plan.entities],
                "parsed_command": nlu_result_for_plan.parsed_command.dict() if nlu_result_for_plan.parsed_command else None, # Add parsed command
                "implicit_goals": None, # Placeholder
                "message": "NLU analysis via NLUProcessor successful."
            }

            # Construct a richer NLU summary for the planning prompt
            nlu_intent_str = first_attempt_nlu_output.get('intent','N/A')
            nlu_score_str = f"{first_attempt_nlu_output.get('intent_score', 0.0):.2f}"
            nlu_entities_list = [e.get('text') for e in first_attempt_nlu_output.get('entities',[])]
            nlu_entities_str = str(nlu_entities_list)[:100] if nlu_entities_list else "None"

            nlu_summary_for_prompt = f"NLU Analysis (NLUProcessor) :: Intent: {nlu_intent_str} (Score: {nlu_score_str}) :: Entities: {nlu_entities_str}"

            if first_attempt_nlu_output.get("parsed_command"):
                cmd = first_attempt_nlu_output["parsed_command"]["command"]
                cmd_params = first_attempt_nlu_output["parsed_command"]["parameters"]
                nlu_summary_for_prompt += f" :: Parsed Command: {cmd} (Params: {str(cmd_params)[:100]})"
            else:
                nlu_summary_for_prompt += " :: Parsed Command: None"

            print(f"{plan_handler_id} INFO: Using NLU result from NLUProcessor: {nlu_summary_for_prompt}")
       else:
            # Fallback to old NLU method if new one failed or is unavailable
            print(f"{plan_handler_id} INFO: NLUProcessor result not available. Falling back to legacy NLU analysis for initial planning.")
            first_attempt_nlu_output = await self.classify_user_intent(user_prompt) # Legacy NLU call
            nlu_summary_for_prompt = f"NLU Analysis (Legacy) :: Intent: {first_attempt_nlu_output.get('intent','N/A')} :: Entities: {str(first_attempt_nlu_output.get('entities',[]))[:100]}..."
            if first_attempt_nlu_output.get("parsed_command"): # Should not happen with legacy, but good for consistency
                cmd = first_attempt_nlu_output["parsed_command"]["command"]
                cmd_params = first_attempt_nlu_output["parsed_command"]["parameters"]
                nlu_summary_for_prompt += f" :: Parsed Command: {cmd} (Params: {str(cmd_params)[:100]})"

            print(f"{plan_handler_id} INFO: Using legacy NLU result: {nlu_summary_for_prompt}")


       while current_attempt <= max_rev_attempts:
           current_attempt_results = [] # Stores results of completed steps for this attempt
           plan_succeeded_this_attempt = True

           # For managing async tasks within the current plan execution attempt
           # Maps step_id to task_id for steps that are currently pending_async
           pending_async_steps: Dict[str, str] = {}
           # Keeps track of step_ids that have been submitted or completed (sync/async)
           # Used to determine if a step has been processed or is waiting for dependencies.
           processed_step_ids_this_attempt = set()


           print(f"{plan_handler_id} Attempt {current_attempt + 1}/{max_rev_attempts + 1}: Generating/Loading plan...")

           # --- Get Contextual Conversation History ---
           # Use configured default values for token and turn limits for context generation
           context_max_tokens = self.conv_ctx_mgr_config["default_max_tokens"]
           context_desired_turns = self.conv_ctx_mgr_config["default_desired_recent_turns"]

           contextual_history_data = await self.conversation_context_manager.get_contextual_history(
               current_prompt_text=user_prompt,
               max_tokens=context_max_tokens,
               desired_recent_turns=context_desired_turns
           )
           history_context_string = self.conversation_context_manager.format_history_for_prompt(contextual_history_data)
           print(f"{plan_handler_id} INFO: Contextual history (est. {contextual_history_data.total_token_estimate} tokens, {contextual_history_data.selected_turns_count} turns) prepared for MasterPlanner using limits: max_tokens={context_max_tokens}, desired_recent_turns={context_desired_turns}.")

           # --- Plan Generation/Loading ---
           agent_capabilities_desc = self.get_agent_capabilities_description()

           kb_query_gen_prompt_str = prompt_constructors.construct_kb_query_generation_prompt(user_prompt, history_context_string, nlu_summary_for_prompt)
           kb_query_agent = next((a for a in self.agents if a.name == "GeneralPurposeAgent"), None) or \
                            next((a for a in self.agents if a.name == "MasterPlanner"), None)
           kb_search_query = ""
           if kb_query_agent:
               kb_query_res = await self._ollama_generate(kb_query_agent.model, kb_query_gen_prompt_str)
               if kb_query_res.get("status") == "success" and kb_query_res.get("response","").strip().upper() != "NO_QUERY_NEEDED":
                   kb_search_query = kb_query_res.get("response","").strip()
                   print(f"{plan_handler_id} INFO: Generated KB search query: '{kb_search_query}'")

           kb_general_ctx_str = await self._get_formatted_general_kb_context(kb_search_query, plan_handler_id)
           kb_plan_log_ctx_str = await self._get_formatted_plan_log_insights(first_attempt_nlu_output, plan_handler_id)
           kb_feedback_ctx_str = await self._get_formatted_feedback_insights(plan_handler_id)

           # --- Query Knowledge Graph for related items based on NLU entities/topics ---
           kg_derived_context_str = ""
           kg_past_plan_summary_context_str = "" # Initialize new context string

           if self.kg_instance and first_attempt_nlu_output.get("status") == "success":
               all_related_kg_items = []
               # Get entities from NLU (for general KG context)
               entities = first_attempt_nlu_output.get("entities", [])
               for entity in entities[:2]: # Limit number of entities to query for KG
                   entity_text_clean = entity.get("text","").strip().lower().replace(' ', '_')
                   if entity_text_clean:
                       # Node ID for keywords (which might represent entities)
                       keyword_node_id = f"keyword_{entity_text_clean}"
                       # Find KB items (sources) that have this keyword (target)
                       related_items = self.kg_instance.get_source_nodes_related_to_target(
                           target_node_id=keyword_node_id,
                           relationship_types=["HAS_KEYWORD"], # Edge: KB_Item -> HAS_KEYWORD -> Keyword_Node
                           limit=2
                       )
                       all_related_kg_items.extend(related_items)

               # Get primary intent as a potential topic to search for
               primary_intent = first_attempt_nlu_output.get("intent")
               if primary_intent:
                   # Standardized node ID format for topics
                   topic_node_id = f"topic_{primary_intent.lower().replace(' ', '_')}"
                   # Find KB items (sources) that have this topic (target)
                   related_items_by_intent_topic = self.kg_instance.get_source_nodes_related_to_target(
                       target_node_id=topic_node_id,
                       relationship_types=["HAS_TOPIC"], # Edge: KB_Item -> HAS_TOPIC -> Topic_Node
                       limit=2
                   )
                   all_related_kg_items.extend(related_items_by_intent_topic)

               # Deduplicate and format for prompt
               unique_related_kg_items_dict = {item['node_id']: item for item in all_related_kg_items}
               unique_related_kg_items = list(unique_related_kg_items_dict.values())

               if unique_related_kg_items:
                   formatted_kg_derived_entries = []
                   for item in unique_related_kg_items[:3]: # Limit total KG derived items in prompt
                       preview = item.get('content_preview', 'N/A')
                       item_type = item.get('node_type', 'Unknown')
                       # 'related_to_target_via' is the key from get_source_nodes_related_to_target
                       related_via = item.get('related_to_target_via', 'unknown relation')
                       formatted_kg_derived_entries.append(f"  - ID: {item['node_id']} (Type: {item_type}, Related to NLU query via: {related_via}): \"{preview[:100]}...\"")
                   if formatted_kg_derived_entries:
                       kg_derived_context_str = "Knowledge Graph Derived Context (KB items related to NLU entities/topics):\n" + "\n".join(formatted_kg_derived_entries) + "\n\n"
                       print(f"[{plan_handler_id}] INFO: Found {len(unique_related_kg_items)} unique KG derived items. Added to context.")
                   else:
                       print(f"[{plan_handler_id}] INFO: No KG derived items formatted for prompt after filtering/processing.")
               else:
                   print(f"[{plan_handler_id}] INFO: No unique KG derived items found from entity/topic graph queries.")

               # Retrieve Simplified Past Plans based on primary intent
               primary_intent = first_attempt_nlu_output.get("intent")
               if primary_intent:
                   topic_node_id = f"topic_{primary_intent.lower().replace(' ', '_')}"
                   # Find SimplifiedPlan nodes (sources) that are related to this intent topic (target)
                   related_simplified_plan_nodes = self.kg_instance.get_source_nodes_related_to_target(
                       target_node_id=topic_node_id,
                       relationship_types=["RELATED_TO_INTENT_TOPIC"], # Edge: SimplifiedPlan -> RELATED_TO_INTENT_TOPIC -> Topic
                       limit=2 # Get 1-2 past plans
                   )
                   if related_simplified_plan_nodes:
                       formatted_past_plans = []
                       for sp_node in related_simplified_plan_nodes:
                           if sp_node.get("node_type") == "SimplifiedPlan":
                               sp_metadata = sp_node.get("metadata")
                               if isinstance(sp_metadata, dict):
                                   sp_json_str = sp_metadata.get("simplified_plan_json")
                                   if sp_json_str and isinstance(sp_json_str, str):
                                       try:
                                           sp_data = SimplifiedPlanStructureDC.from_json_string(sp_json_str)
                                           status_emoji = "✅" if sp_data.status == "success" else "❌"
                                           formatted_past_plans.append(
                                               f"  - Past Plan (Intent: {sp_data.primary_intent}, Status: {sp_data.status} {status_emoji}, Agents: [{', '.join(sp_data.agent_sequence[:3])}...]): Preview: '{sp_data.original_request_preview[:50]}...'"
                                           )
                                       except Exception as e_sp_parse:
                                           print(f"[{plan_handler_id}] WARNING: Failed to parse SimplifiedPlanStructureDC from KG node {sp_node.get('node_id')}'s metadata: {e_sp_parse}")
                                   else:
                                       print(f"[{plan_handler_id}] WARNING: 'simplified_plan_json' not found or not a string in metadata for SimplifiedPlan node {sp_node.get('node_id')}.")
                               else:
                                   print(f"[{plan_handler_id}] WARNING: Metadata for SimplifiedPlan node {sp_node.get('node_id')} is not a dict.")
                       if formatted_past_plans:
                           kg_past_plan_summary_context_str = "Past Simplified Plan Structures (for similar intent):\n" + "\n".join(formatted_past_plans) + "\n\n"
                           print(f"[{plan_handler_id}] INFO: Retrieved {len(formatted_past_plans)} simplified past plans from KG.")
                       else:
                            print(f"[{plan_handler_id}] INFO: No suitable simplified past plans found or parsed from KG for intent '{primary_intent}'.")
                   else:
                       print(f"[{plan_handler_id}] INFO: No simplified past plans linked to topic '{topic_node_id}' found in KG.")
           else: # KG instance not available or NLU failed
               print(f"[{plan_handler_id}] INFO: KG instance not available or NLU failed, skipping KG derived context and past plan retrieval.")

           current_rl_state_kb_summary = {
               "general_hits_count": len(kb_general_ctx_str.splitlines()) -1 if kb_general_ctx_str.strip() else 0,
               "kg_derived_hits_count": len(kg_derived_context_str.splitlines()) -1 if kg_derived_context_str.strip() else 0, # Retained from previous step
               "past_plan_summary_hits_count": len(kg_past_plan_summary_context_str.splitlines()) -1 if kg_past_plan_summary_context_str.strip() else 0, # New
               "plan_log_hits_count": len(kb_plan_log_ctx_str.splitlines()) -1 if kb_plan_log_ctx_str.strip() else 0,
               "feedback_hits_count": len(kb_feedback_ctx_str.splitlines()) -1 if kb_feedback_ctx_str.strip() else 0,
               "kg_derived_hits_count": len(kg_derived_context_str.splitlines()) -1 if kg_derived_context_str.strip() else 0,
               "past_plan_summary_hits_count": len(kg_past_plan_summary_context_str.splitlines()) -1 if kg_past_plan_summary_context_str.strip() else 0, # New
           }
           current_rl_state = self._construct_rl_state(user_prompt, first_attempt_nlu_output, current_rl_state_kb_summary, None)

           # --- RL-based Planner Strategy Selection ---
           # Define available conceptual strategies for the MasterPlanner's prompting approach.
           available_planner_strategies = ["Strategy_Default", "Strategy_FocusClarity", "Strategy_PrioritizeBrevity"]
           selected_strategy = "Strategy_Default" # Default strategy if RL cannot choose.

           # Ensure planner_agent is resolved to get its model details for logging, if available.
           planner_agent = next((a for a in self.agents if a.name == "MasterPlanner" and a.active), None)

           if self.rl_policy_manager and planner_agent:
               # Construct the state key from the current RL state dictionary.
               state_key_for_rl = self.rl_policy_manager._construct_state_key(current_rl_state)
               # Ask the RL policy manager for the best known action (strategy) for this state.
               chosen_action = self.rl_policy_manager.get_best_action(state_key_for_rl, available_planner_strategies)

               if chosen_action:
                   selected_strategy = chosen_action
                   print(f"[{plan_handler_id}] RLPolicyManager selected strategy: {selected_strategy} for state: {state_key_for_rl}")
               else:
                   # No specific preference from RL manager, or exploration led to random choice (handled in get_best_action).
                   # If get_best_action returned None (e.g. no actions available), selected_strategy remains default.
                   print(f"[{plan_handler_id}] RLPolicyManager provided no overriding strategy or chose default for state: {state_key_for_rl}. Using: {selected_strategy}")
           elif not planner_agent:
                print(f"[{plan_handler_id}] MasterPlanner agent not found. Using default strategy: {selected_strategy}.")
           else: # RL Policy Manager not available (e.g., failed to initialize)
               print(f"[{plan_handler_id}] RLPolicyManager not available. Using default strategy: {selected_strategy}.")

           # This 'selected_strategy' will be logged as the 'action_taken' for this RL experience.
           current_rl_action = selected_strategy
           # Log details about the prompt generation context, including the chosen strategy.
           current_prompt_details = {
               "strategy_used": selected_strategy,
               "llm_model": planner_agent.model if planner_agent else "MasterPlanner_Agent_Not_Found"
            }
           # NOTE: For this initial RL integration, the actual content of 'planning_prompt' below
           # does NOT yet change based on 'selected_strategy'. The learning loop works by associating
           # the *name* of the strategy (logged in 'action_taken') with the eventual reward.
           # Future enhancements would involve `construct_main_planning_prompt` generating
           # different prompt structures based on the value of `selected_strategy`.

           if state_for_executed_plan_log is None : # This captures details of the first planning attempt for logging.
                state_for_executed_plan_log = current_rl_state
                action_for_executed_plan_log = current_rl_action
                prompt_details_for_executed_plan_log = current_prompt_details

           if current_attempt == 0:
               # The actual planning prompt construction happens here.
               # Pass the selected_strategy to the prompt constructor.
               planning_prompt = prompt_constructors.construct_main_planning_prompt(
                   user_prompt=user_prompt,
                   history_context=history_context_string,
                   nlu_summary=nlu_summary_for_prompt,
                   kb_general_context=kb_general_ctx_str,
                   kg_derived_context=kg_derived_context_str,
                   kg_past_plan_summary_context=kg_past_plan_summary_context_str,
                   kb_plan_log_context=kb_plan_log_ctx_str,
                   kb_feedback_context=kb_feedback_ctx_str,
                   agent_capabilities_description=agent_capabilities_desc,
                   planner_strategy=selected_strategy # Pass the chosen strategy
               )
           else: # Constructing a revision prompt
               print(f"{plan_handler_id} INFO: Constructing revision prompt with failure context.")
               planning_prompt = prompt_constructors.construct_revision_planning_prompt(
                   user_prompt=user_prompt,
                   history_context=history_context_string, # Use formatted string
                   nlu_summary=nlu_summary_for_prompt,
                   failure_context=detailed_failure_ctx_for_rev,
                   agent_capabilities_description=agent_capabilities_desc
               )

           planner_agent = next((a for a in self.agents if a.name == "MasterPlanner" and a.active), None)
           if not planner_agent: return [{"status":"error", "message":"MasterPlanner agent not found/active."}]
           raw_plan_response = await self._ollama_generate(planner_agent.model, planning_prompt)
           original_plan_json_str = raw_plan_response.get("response", "[]")

           try:
               plan_list = json.loads(original_plan_json_str)
               if not isinstance(plan_list, list): raise ValueError("Plan is not a list.")
               for step in plan_list:
                   if not step.get("step_type") in ["conditional", "loop", "parallel_group", "agent_service_call"] and \
                      not all(k in step for k in ["step_id", "agent_name", "task_prompt"]):
                       raise ValueError(f"Regular agent step missing required keys: {step}")
           except Exception as e_parse:
               print(f"{plan_handler_id} ERROR: Failed to parse plan from LLM: {e_parse}. Response: {original_plan_json_str[:200]}...")
               if current_attempt < max_rev_attempts:
                   detailed_failure_ctx_for_rev = self._capture_simple_failure_context(original_plan_json_str, {"error": f"Plan parsing failed: {e_parse}"})
                   current_attempt += 1; step_outputs = {}; plan_succeeded_this_attempt = False
                   state_for_executed_plan_log = current_rl_state
                   action_for_executed_plan_log = current_rl_action
                   prompt_details_for_executed_plan_log = current_prompt_details
                   continue
               else:
                   final_exec_results.append({"status":"error", "message":f"Failed to parse plan after {max_rev_attempts+1} attempts. Last error: {e_parse}"})
                   plan_succeeded_this_attempt = False; break

           if not plan_list:
               print(f"{plan_handler_id} INFO: Planner returned an empty plan.");
               plan_succeeded_this_attempt = True; # No steps, so technically success.
               final_exec_results = [] # Ensure it's defined
               break # Exit revision loop if plan is empty

           # --- Main Plan Execution Loop with Async Task Management ---
           # executed_step_ids tracks steps whose *final result* (sync or async) is processed.
           executed_step_ids = set()
           active_loops = {} # For managing loop contexts (as before)
           loop_context_stack = [] # For managing nested loops (as before)

           # --- Main Plan Execution Loop with Async Task Management ---
           # This loop continues as long as there are steps in the plan that haven't been successfully executed,
           # OR if there are asynchronous tasks still pending completion for the current plan attempt.
           #
           # The loop has three main parts in each iteration:
           # 1. Dispatch Ready Steps: Iterate through the plan_list. If a step's dependencies are met
           #    (i.e., all preceding steps it depends on are in `executed_step_ids`) and it's not already
           #    pending or executed, then dispatch it.
           #    - If a step is dispatched and completes synchronously (success or failure), its result is recorded,
           #      and it's added to `executed_step_ids`.
           #    - If a step is dispatched and returns a `pending_async` status (with a `task_id`), it's added
           #      to the `pending_async_steps` dictionary (mapping step_id to task_id).
           # 2. Poll Pending Async Tasks: Iterate through `pending_async_steps`. For each, check its status
           #    using `get_async_task_info()`.
           #    - If COMPLETED: Process its result, update `step_outputs`, add to `current_attempt_results`
           #      and `executed_step_ids`. Remove from `pending_async_steps`.
           #    - If FAILED: Record failure, add to `executed_step_ids`, remove from `pending_async_steps`,
           #      and mark the overall plan attempt as failed.
           # 3. Sleep if Idle: If no new steps were dispatched in part 1, and no async tasks were resolved in part 2,
           #    but there are still tasks in `pending_async_steps`, then sleep briefly to avoid busy-waiting
           #    before the next polling cycle.
           #
           # The loop terminates for the current attempt if:
           #    a) All steps in `plan_list` are in `executed_step_ids` AND `pending_async_steps` is empty (plan success).
           #    b) `plan_succeeded_this_attempt` becomes `False` (due to a step failure, sync or async).
           #    c) A deadlock/stall is detected (no progress, no pending tasks, but plan not complete).
           #
           # executed_step_ids tracks steps whose *final result* (sync or async) is processed.
           executed_step_ids = set()
           active_loops = {} # For managing loop contexts (as before)
           loop_context_stack = [] # For managing nested loops (as before)

           while len(executed_step_ids) < len(plan_list) or pending_async_steps:
               dispatched_this_cycle = False # Track if any new step was dispatched or async task completed

               # 1. Attempt to dispatch new steps
               current_step_idx_for_dispatch = 0
               while current_step_idx_for_dispatch < len(plan_list):
                   step_to_evaluate = plan_list[current_step_idx_for_dispatch]
                   step_id_to_evaluate = step_to_evaluate.get("step_id")

                   if step_id_to_evaluate in executed_step_ids or step_id_to_evaluate in pending_async_steps:
                       current_step_idx_for_dispatch +=1; continue # Already done or pending

                   # Check dependencies: all dependent step_ids must be in executed_step_ids
                   can_execute = True
                   for dep_id in step_to_evaluate.get("dependencies",[]):
                       if dep_id not in executed_step_ids:
                           # print(f"{plan_handler_id} Deferring step {step_id_to_evaluate}: Unmet dependency {dep_id}")
                           can_execute = False; break

                   if not can_execute:
                       current_step_idx_for_dispatch +=1; continue

                   # --- Dispatching the step ---
                   dispatched_this_cycle = True # A step is being dispatched
                   processed_step_ids_this_attempt.add(step_id_to_evaluate)

                   step_priority = step_to_evaluate.get("priority", "normal").lower()
                   dispatch_log_prefix = f"[{plan_handler_id}]" # Re-init for each step for clarity
                   if step_priority == "high": dispatch_log_prefix += f" [Priority: HIGH]"

                   step_type_for_log = step_to_evaluate.get("step_type", "agent_execution")
                   log_agent_name = step_to_evaluate.get('agent_name', step_to_evaluate.get('target_agent_name', 'N/A'))
                   log_desc = step_to_evaluate.get('description', step_to_evaluate.get('service_name', 'N/A'))[:50]
                   print(f"{dispatch_log_prefix} Dispatching Step {step_id_to_evaluate}: Type='{step_type_for_log}', Agent/Service='{log_agent_name}', Desc='{log_desc}...'")

                   step_result_or_task: Dict # To hold result from sync step or task info from async submission
                   step_type = step_to_evaluate.get("step_type", "agent_execution")
                   agent_name_for_step = step_to_evaluate.get("agent_name")

                   # --- Step Type Handling (Conditional, Loop, Service Call, Regular Agent, Tool Suggestion) ---
                   if agent_name_for_step == "SystemCapabilityManager" and step_to_evaluate.get("task_prompt") == "SUGGEST_NEW_TOOL":
                       tool_desc = step_to_evaluate.get("suggested_tool_description", "No description provided.")
                       log_message = f"{datetime.datetime.utcnow().isoformat()} - Tool Suggestion by MasterPlanner: {tool_desc}\n"
                       try:
                           with open(self.logs_dir / "tool_suggestions.log", "a", encoding="utf-8") as f:
                               f.write(log_message)
                           step_result_or_task = {"step_id": step_id_to_evaluate, "agent_name": "SystemCapabilityManager", "status": "success", "response": "Tool suggestion logged successfully."}
                           print(f"{dispatch_log_prefix} Logged tool suggestion: {tool_desc}")
                       except Exception as e_log:
                           print(f"{dispatch_log_prefix} ERROR logging tool suggestion: {e_log}")
                           step_result_or_task = {"step_id": step_id_to_evaluate, "agent_name": "SystemCapabilityManager", "status": "error", "response": f"Failed to log tool suggestion: {e_log}"}

                   elif step_type == "conditional":
                       next_step_id_from_cond, eval_res_cond = await self._handle_conditional_step(step_to_evaluate, plan_list, step_outputs, executed_step_ids, plan_handler_id)
                       step_result_or_task = eval_res_cond if eval_res_cond else {"step_id": step_id_to_evaluate, "status":"error", "response":"Conditional eval result missing"}
                       # Conditional step itself is now "executed" (its evaluation is complete)
                       executed_step_ids.add(step_id_to_evaluate)
                       current_attempt_results.append(step_result_or_task) # Log its evaluation result
                       if step_result_or_task.get("status") != "success": plan_succeeded_this_attempt = False; break
                       # Note: Jump logic for conditionals is implicitly handled by the dispatcher finding the next valid step based on dependencies and executed_ids.

                   elif step_type == "loop" and step_to_evaluate.get("loop_type") == "while":
                       # TODO: Loop logic needs significant review for full async compatibility within loop bodies.
                       # Current _handle_loop_step might not integrate perfectly if loop body steps become async.
                       # For now, treating loop header evaluation as synchronous.
                       print(f"{plan_handler_id} DEBUG: Loop step {step_id_to_evaluate} encountered. Advanced async loop body handling is TBD.")
                       # Simplified: Assume loop header itself is a synchronous operation for now.
                       # A more robust solution would involve making _handle_loop_step async-aware or
                       # integrating loop flow control directly into this main execution dispatcher.
                       # This is a placeholder to keep the structure.
                       step_result_or_task = {"status":"success", "response": f"Loop header {step_id_to_evaluate} processed (simulated)."}
                       executed_step_ids.add(step_id_to_evaluate) # Mark loop header as executed
                       current_attempt_results.append(step_result_or_task)

                   elif step_type == "agent_service_call":
                       step_result_or_task = await self._handle_agent_service_call(step_to_evaluate, step_outputs, plan_list)

                   elif step_to_evaluate.get("agent_name") == "parallel_group": # Not fully implemented yet
                       # TODO: Parallel group execution would involve submitting multiple sub-steps as async tasks
                       # and then waiting for all of them, potentially using asyncio.gather.
                       print(f"{dispatch_log_prefix} DEBUG: Parallel group {step_id_to_evaluate} encountered (simulated as synchronous success).")
                       step_result_or_task = {"status": "success", "response": f"Parallel group {step_id_to_evaluate} processed (simulated)."}

                   else: # Regular agent execution step (e.g., direct agent call, not a service)
                       step_result_or_task = await self._execute_single_plan_step(step_to_evaluate, plan_list, step_outputs)

                   # --- Process result of the dispatched step ---
                   if step_result_or_task.get("status") == "pending_async":
                       # The step initiated an asynchronous task.
                       task_id = step_result_or_task["task_id"]
                       pending_async_steps[step_id_to_evaluate] = task_id # Map step_id to task_id
                       print(f"{dispatch_log_prefix} Step {step_id_to_evaluate} initiated as PENDING_ASYNC task_id: {task_id}.")

                   elif step_result_or_task.get("status") == "success":
                       # The step completed synchronously and successfully.
                       current_attempt_results.append(step_result_or_task)
                       executed_step_ids.add(step_id_to_evaluate) # Mark as fully completed
                       print(f"{dispatch_log_prefix} Step {step_id_to_evaluate} completed synchronously with SUCCESS.")

                   else: # The step completed synchronously but failed.
                       current_attempt_results.append(step_result_or_task)
                       executed_step_ids.add(step_id_to_evaluate) # Mark as completed (albeit failed)
                       plan_succeeded_this_attempt = False # Mark entire plan attempt as failed
                       print(f"{dispatch_log_prefix} Step {step_id_to_evaluate} completed synchronously with FAILURE: {step_result_or_task.get('response')}")
                       break # Exit the dispatch loop for this attempt, as a step has failed.

                   current_step_idx_for_dispatch +=1
               # End of dispatch loop (part 1 of main while loop)

               if not plan_succeeded_this_attempt: break # Exit main execution while loop for this attempt if a sync step failed.

               # 2. Check status of pending asynchronous tasks
               if pending_async_steps:
                   completed_tasks_this_cycle: List[str] = [] # step_ids of tasks that finished this cycle
                   for step_id_pending, task_id_pending in list(pending_async_steps.items()): # list() for safe removal
                       task_info = await self.get_async_task_info(task_id_pending)

                       if task_info: # Ensure task info was found
                           if task_info.status == AsyncTaskStatus.COMPLETED:
                               dispatched_this_cycle = True # Activity occurred (an async task completed)
                               print(f"{plan_handler_id} AsyncTask {task_id_pending} for step {step_id_pending} has COMPLETED.")

                               async_step_result = task_info.result # This is the dict from execute_agent or service call
                               if not isinstance(async_step_result, dict):
                                   # This indicates an issue with what the async task's coroutine returned to the wrapper.
                                   print(f"{plan_handler_id} ERROR: Async task {task_id_pending} (step {step_id_pending}) result is not a dict: {async_step_result}")
                                   async_step_result = {"status": "error", "response": f"Async task result format error for step {step_id_pending}", "data": None}

                               # Store output if successful and output_variable_name is defined
                               original_step_def_for_async = next((s_def for s_def in plan_list if s_def.get("step_id") == step_id_pending), None)
                               output_var_name_for_async = original_step_def_for_async.get("output_variable_name") if original_step_def_for_async else None

                               if async_step_result.get("status") == "success" and output_var_name_for_async:
                                   data_to_store = async_step_result.get("data", async_step_result.get("response"))
                                   step_outputs[output_var_name_for_async] = data_to_store
                                   # Handle other special keys if present in the async result (e.g., image_path)
                                   for mk_async in ["image_path","frame_path","gif_path","speech_path","modified_file"]:
                                       if mk_async in async_step_result: step_outputs[f"{output_var_name_for_async}_{mk_async}"]=async_step_result[mk_async]

                               # Log the result of the completed async step
                               agent_name_for_async_log = "UnknownAgent"
                               if original_step_def_for_async:
                                   agent_name_for_async_log = original_step_def_for_async.get('agent_name',
                                                               original_step_def_for_async.get('target_agent_name', 'AsyncStep'))
                               current_attempt_results.append({
                                   "step_id": step_id_pending,
                                   "agent_name": agent_name_for_async_log,
                                   "status": async_step_result.get("status"),
                                   "response": async_step_result.get("message", async_step_result.get("response")),
                                   "data": async_step_result.get("data")
                                })
                                executed_step_ids.add(step_id_pending) # Mark step as fully executed
                                completed_tasks_this_cycle.append(step_id_pending)
                                if async_step_result.get("status") != "success":
                                    plan_succeeded_this_attempt = False # Mark plan attempt as failed

                           elif task_info.status == AsyncTaskStatus.FAILED:
                               dispatched_this_cycle = True # Activity occurred
                               print(f"{plan_handler_id} AsyncTask {task_id_pending} for step {step_id_pending} has FAILED: {task_info.error}")
                               current_attempt_results.append({"step_id": step_id_pending, "status": "error", "response": task_info.error, "agent_name": "AsyncTaskManager"})
                               executed_step_ids.add(step_id_pending) # Mark as executed (failed)
                               completed_tasks_this_cycle.append(step_id_pending)
                               plan_succeeded_this_attempt = False # Mark plan attempt as failed

                           # If PENDING or RUNNING, do nothing this cycle for this task, wait for next poll.

                       else: # Task info not found - should be rare if IDs are managed correctly
                           print(f"{plan_handler_id} ERROR: No task info found for pending task_id {task_id_pending} (step {step_id_pending}). Marking step as error.")
                           current_attempt_results.append({"step_id": step_id_pending, "status": "error", "response": "Async task info lost.", "agent_name": "AsyncTaskManager"})
                           executed_step_ids.add(step_id_pending)
                           completed_tasks_this_cycle.append(step_id_pending)
                           plan_succeeded_this_attempt = False

                   # Remove tasks that completed or failed this cycle from pending_async_steps
                   for step_id_done in completed_tasks_this_cycle:
                       pending_async_steps.pop(step_id_done, None)

                   if not plan_succeeded_this_attempt: break # Exit main execution while loop for this attempt if an async task failed.

               # 3. Sleep if no activity but tasks are still pending, or break if deadlocked
               if not dispatched_this_cycle and pending_async_steps:
                   # No new steps were ready to dispatch, and no pending tasks finished this cycle.
                   # Sleep briefly before polling pending tasks again.
                   print(f"{plan_handler_id} No new dispatches or task completions this cycle. {len(pending_async_steps)} tasks still pending. Sleeping...")
                   await asyncio.sleep(0.2)
               elif not dispatched_this_cycle and not pending_async_steps and len(executed_step_ids) < len(plan_list):
                   # No new dispatches, no pending tasks, but plan is not fully executed.
                   # This indicates a potential deadlock (e.g., circular dependencies or unmet dependencies).
                   print(f"{plan_handler_id} WARNING: No progress made (no new dispatches, no pending tasks), but plan not complete ({len(executed_step_ids)}/{len(plan_list)} steps executed). Check for dependency errors or logic issues in plan.")
                   plan_succeeded_this_attempt = False # Consider this a failure of the plan execution.
                   break # Exit main execution loop for this attempt.

           # --- End of Main Plan Execution Loop for this attempt ---
           final_exec_results = current_attempt_results # These are the results of all processed steps for this attempt

           if not plan_succeeded_this_attempt and current_attempt < max_rev_attempts:
               # Capture failure context for revision. If failure was due to an async step, find its details in current_attempt_results.
               failed_step_details_for_context = next((res for res in reversed(current_attempt_results) if res.get("status") != "success"), None)
               failed_step_def_for_context = None
               if failed_step_details_for_context and failed_step_details_for_context.get("step_id"):
                   failed_step_def_for_context = next((s_def for s_def in plan_list if s_def.get("step_id") == failed_step_details_for_context.get("step_id")), None)

               detailed_failure_ctx_for_rev = self._capture_failure_context(
                   original_plan_json_str,
                   failed_step_def_for_context,
                   failed_step_details_for_context,
                   step_outputs
               )
               current_attempt += 1;
               step_outputs = {}; # Reset outputs for next attempt
               # Reset tracking for next attempt
               pending_async_steps.clear()
               processed_step_ids_this_attempt.clear()
               executed_step_ids.clear()
               state_for_executed_plan_log = current_rl_state
               action_for_executed_plan_log = current_rl_action
               prompt_details_for_executed_plan_log = current_prompt_details
           else: break

       user_facing_summary = await self._summarize_execution_for_user(user_prompt, final_exec_results)
       final_plan_outcome_status_str = "success" if plan_succeeded_this_attempt else "failure"
       user_feedback_rating_for_log = "none"
       if state_for_executed_plan_log is None: state_for_executed_plan_log = self._construct_rl_state(user_prompt, first_attempt_nlu_output, {}, None)
       if action_for_executed_plan_log is None: action_for_executed_plan_log = "Action_Planner_Error_Before_Action"
       if prompt_details_for_executed_plan_log is None: prompt_details_for_executed_plan_log = {"error": "No plan generated or parsed"}
       calculated_reward = self._calculate_rl_reward(final_plan_outcome_status_str, user_feedback_rating_for_log, current_attempt)
       timestamp_interaction_end = datetime.datetime.now().isoformat()
       self.rl_logger.log_experience(
           rl_interaction_id=rl_interaction_id, attempt_number=current_attempt, state=state_for_executed_plan_log, action=action_for_executed_plan_log,
           master_planner_prompt_details=prompt_details_for_executed_plan_log, generated_plan_json=original_plan_json_str,
           plan_parsing_status="success" if plan_list else "failure",
           final_executed_plan_json=original_plan_json_str if plan_succeeded_this_attempt or (not plan_succeeded_this_attempt and current_attempt >=max_rev_attempts) else "N/A",
           execution_status=final_plan_outcome_status_str, user_feedback_rating=user_feedback_rating_for_log,
           calculated_reward=calculated_reward, next_state=None, done=True,
           timestamp_start_iso=timestamp_interaction_start, timestamp_end_iso=timestamp_interaction_end
       )
       # Publish an event to notify listeners (like RLPolicyManager) that a new RL experience has been logged
       # and policy updates might be needed.
       await self.publish_event(
           event_type="rl.experience.logged",
           source_component="TerminusOrchestrator.RLLogger",
           payload={
               "log_file_path": str(self.rl_logger.log_file_path),
               "rl_interaction_id": rl_interaction_id,
               "execution_status": final_plan_outcome_status_str
            }
       )
       print(f"[{plan_handler_id}] Published 'rl.experience.logged' event for interaction {rl_interaction_id}.")

       if plan_list or not plan_succeeded_this_attempt :
            current_plan_log_kb_id = await self._store_plan_execution_log_in_kb(user_prompt, first_attempt_nlu_output, original_plan_json_str, plan_succeeded_this_attempt, current_attempt + 1, final_exec_results, step_outputs, user_facing_summary)

       # Add assistant's summary to history as a ConversationTurn object
       assistant_turn_metadata = {
           "is_plan_outcome": True,
           "plan_log_kb_id": current_plan_log_kb_id,
           "feedback_item_id": current_plan_log_kb_id, # Assuming feedback can be tied to the plan log
           "feedback_item_type": "master_plan_log_outcome",
           "related_user_prompt_for_feedback": user_prompt
       }
       assistant_turn = ConversationTurn(role="assistant", content=user_facing_summary, metadata=assistant_turn_metadata)
       self.conversation_history.append(assistant_turn)
       # Prune again if history exceeds max items after adding assistant turn
       if len(self.conversation_history) > self.max_history_items:
           self.conversation_history = self.conversation_history[-self.max_history_items:]
       # Update context manager after assistant's turn as well
       self.conversation_context_manager.update_full_history(self.conversation_history)

       return final_exec_results

   async def _evaluate_plan_condition(self, condition_def: Dict, step_outputs: Dict, full_plan_list: List[Dict]) -> Dict:
       if not condition_def: return {"status": "error", "evaluation": None, "message": "Condition definition is empty."}
       source_step_id = condition_def.get("source_step_id"); output_var_path = condition_def.get("source_output_variable"); operator = condition_def.get("operator"); compare_value_literal = condition_def.get("value"); value_type_hint = condition_def.get("value_type", "string")
       if not all([source_step_id, output_var_path, operator]): return {"status": "error", "evaluation": None, "message": "Condition missing source_step_id, source_output_variable, or operator."}
       source_step_output_key = None
       for step_cfg in full_plan_list:
           if step_cfg.get("step_id") == source_step_id: source_step_output_key = step_cfg.get("output_variable_name", f"step_{source_step_id}_output"); break
       if not source_step_output_key or source_step_output_key not in step_outputs: return {"status": "error", "evaluation": None, "message": f"Output for source_step_id '{source_step_id}' (expected key: {source_step_output_key}) not found in step_outputs: {list(step_outputs.keys())}."}
       actual_value_raw = step_outputs.get(source_step_output_key)
       try:
           current_val = actual_value_raw
           for part in output_var_path.split('.'):
               if isinstance(current_val, dict): current_val = current_val.get(part);
               elif isinstance(current_val, list) and part.isdigit() and int(part) < len(current_val): current_val = current_val[int(part)]
               else: current_val = None; break
           actual_value_to_compare = current_val
       except Exception as e: return {"status": "error", "evaluation": None, "message": f"Error accessing source_output_variable '{output_var_path}' from '{source_step_output_key}': {str(e)}"}
       try:
           if value_type_hint == "integer": actual_value_to_compare = int(actual_value_to_compare) if actual_value_to_compare is not None else None; compare_value_literal = int(compare_value_literal) if compare_value_literal is not None else None
           elif value_type_hint == "float": actual_value_to_compare = float(actual_value_to_compare) if actual_value_to_compare is not None else None; compare_value_literal = float(compare_value_literal) if compare_value_literal is not None else None
           elif value_type_hint == "boolean":
               if isinstance(actual_value_to_compare, str): actual_value_to_compare = actual_value_to_compare.lower() in ["true", "1", "yes"]
               else: actual_value_to_compare = bool(actual_value_to_compare)
               if isinstance(compare_value_literal, str): compare_value_literal = compare_value_literal.lower() in ["true", "1", "yes"]
               else: compare_value_literal = bool(compare_value_literal)
           else: actual_value_to_compare = str(actual_value_to_compare) if actual_value_to_compare is not None else ""; compare_value_literal = str(compare_value_literal) if compare_value_literal is not None else ""
       except ValueError as ve: return {"status": "error", "evaluation": None, "message": f"Type conversion error for '{value_type_hint}': {ve}. Val: '{actual_value_to_compare}', Comp: '{compare_value_literal}'"}
       evaluation = None
       try:
           if operator == "equals": evaluation = (actual_value_to_compare == compare_value_literal)
           elif operator == "not_equals": evaluation = (actual_value_to_compare != compare_value_literal)
           elif operator == "contains":
               if isinstance(actual_value_to_compare, (str, list)): evaluation = (compare_value_literal in actual_value_to_compare)
               elif isinstance(actual_value_to_compare, dict): evaluation = (compare_value_literal in actual_value_to_compare.keys() or compare_value_literal in actual_value_to_compare.values())
               else: return {"status": "error", "evaluation": None, "message": f"'contains' needs str/list/dict; got {type(actual_value_to_compare)}"}
           elif operator == "not_contains":
               if isinstance(actual_value_to_compare, (str, list)): evaluation = (compare_value_literal not in actual_value_to_compare)
               elif isinstance(actual_value_to_compare, dict): evaluation = not (compare_value_literal in actual_value_to_compare.keys() or compare_value_literal in actual_value_to_compare.values())
               else: return {"status": "error", "evaluation": None, "message": f"'not_contains' needs str/list/dict; got {type(actual_value_to_compare)}"}
           elif operator == "is_true": evaluation = bool(actual_value_to_compare)
           elif operator == "is_false": evaluation = not bool(actual_value_to_compare)
           elif operator == "is_empty":
               if hasattr(actual_value_to_compare, '__len__'): evaluation = (len(actual_value_to_compare) == 0)
               else: evaluation = (actual_value_to_compare is None)
           elif operator == "is_not_empty":
               if hasattr(actual_value_to_compare, '__len__'): evaluation = (len(actual_value_to_compare) > 0)
               else: evaluation = (actual_value_to_compare is not None)
           elif operator == "greater_than":
               if isinstance(actual_value_to_compare, (int, float)) and isinstance(compare_value_literal, (int, float)): evaluation = (actual_value_to_compare > compare_value_literal)
               else: return {"status": "error", "evaluation": None, "message": "Numeric types required for 'greater_than'."}
           elif operator == "less_than":
               if isinstance(actual_value_to_compare, (int, float)) and isinstance(compare_value_literal, (int, float)): evaluation = (actual_value_to_compare < compare_value_literal)
               else: return {"status": "error", "evaluation": None, "message": "Numeric types required for 'less_than'."}
           else: return {"status": "error", "evaluation": None, "message": f"Unsupported operator: {operator}"}
           return {"status": "success", "evaluation": evaluation, "message": f"Condition '{output_var_path} {operator} {compare_value_literal}' (Actual: {actual_value_to_compare}) evaluated to {evaluation}."}
       except Exception as e: return {"status": "error", "evaluation": None, "message": f"Error during condition evaluation: {str(e)}"}

   def _capture_failure_context(self, plan_str:str, failed_step_def:Optional[Dict], failed_step_result:Optional[Dict], current_outputs:Dict) -> Dict:
        return {"plan_that_failed_this_attempt": plan_str, "failed_step_definition": failed_step_def if failed_step_def else "N/A", "failed_step_execution_result": failed_step_result if failed_step_result else "N/A", "step_outputs_before_failure": dict(current_outputs)}

   async def _summarize_execution_for_user(self, user_prompt:str, final_exec_results:List[Dict]) -> str:
       summarizer = next((a for a in self.agents if a.name=="CreativeWriter"),None) or next((a for a in self.agents if a.name=="DeepThink"),None)
       if not summarizer: s_count = sum(1 for r in final_exec_results if r.get('status')=='success'); return f"Plan execution finished. {s_count}/{len(final_exec_results)} steps processed successfully."
       summary_context = f"Original User Request: '{user_prompt}'\nPlan Execution Summary of Final Attempt:\n"
       if not final_exec_results: summary_context += "No steps were executed or the plan was empty.\n"
       else:
           for i, res in enumerate(final_exec_results): summary_context += (f"  Step {i+1} (Agent: {res.get('agent','N/A')}, ID: {res.get('step_id','N/A')}): Status='{res.get('status','unknown')}', Output Snippet='{str(res.get('response','No response'))[:100]}...'\n")
       prompt = (f"You are an AI assistant. Based on the user's original request and a summary of the multi-step plan execution's final attempt, provide a concise, natural language summary of what actions were taken by the system and the overall outcome. Focus on what would be most useful for the user to understand what just happened. Do not refer to yourself as '{summarizer.name}', just act as the main AI assistant.\n\n{summary_context}\n\nPlease provide only the summary text, suitable for conversation history.")
       res = await self.execute_agent(summarizer, prompt)
       if res.get("status")=="success" and res.get("response","").strip(): return res.get("response").strip()
       else: s_count = sum(1 for r in final_exec_results if r.get('status')=='success'); return f"Plan execution attempt finished. {s_count}/{len(final_exec_results)} steps successful. Summarization failed: {res.get('response')}"

   async def _store_plan_execution_log_in_kb(self, user_prompt_orig:str, nlu_output_orig:Dict, plan_json_final_attempt:str, final_status_bool:bool, num_attempts:int, step_results_final_attempt:List[Dict], outputs_final_attempt:Dict, user_facing_summary_text:str) -> Optional[str]:
       if not self.knowledge_collection:
           print("MasterPlanner: Knowledge Base unavailable, skipping storage of plan execution summary.");
           return None

       final_plan_status_str = "success" if final_status_bool else "failure"

       summary_list_for_log = [
           {"step_id": s.get("step_id", "N/A"),
            "agent_name": s.get("agent_name", s.get("agent", "N/A")), # Use agent_name from new structure
            "status": s.get("status", "unknown"),
            "response_preview": str(s.get("response", ""))[:150] + "..."
           } for s in step_results_final_attempt
       ]

       nlu_analysis_data_for_log = {}
       if isinstance(nlu_output_orig, dict):
           nlu_analysis_data_for_log = {
               "intent": nlu_output_orig.get("intent"),
               "intent_scores": nlu_output_orig.get("intent_scores"),
               "entities": nlu_output_orig.get("entities", [])
           }

       # Create PlanExecutionRecordDC instance
       plan_record = PlanExecutionRecordDC(
           # record_id and timestamp_utc are auto-generated
           original_user_request=user_prompt_orig,
           primary_intent=nlu_analysis_data_for_log.get("intent"),
           nlu_analysis_raw=nlu_analysis_data_for_log if nlu_analysis_data_for_log else None, # Store the whole NLU dict
           status=final_plan_status_str,
           total_attempts=num_attempts,
           plan_json_executed_final_attempt=plan_json_final_attempt, # This is already a JSON string
           final_summary_to_user=user_facing_summary_text,
           step_results_summary=summary_list_for_log,
           final_step_outputs={k: (str(v)[:200]+"..." if len(str(v)) > 200 else v) for k,v in outputs_final_attempt.items()},
           rl_interaction_id=getattr(self, 'rl_interaction_id', None) # Assuming rl_interaction_id is available on self or passed
                                                                    # For now, this might be None if not set up in execute_master_plan context
                                                                    # This was just an example, rl_interaction_id is set in execute_master_plan
                                                                    # Let's assume it's available via a property or passed if needed.
                                                                    # For this refactor, let's make it simple and pass it if available, else None.
                                                                    # Actually, rl_interaction_id is defined in execute_master_plan, not directly here.
                                                                    # For now, this field can be None or we can pass rl_interaction_id to this func.
                                                                    # Let's assume it's not critical for the schema structure itself for now.
       )

       # Metadata for top-level ChromaDB filtering (complementing the structured data)
       kb_meta_for_filtering = {
           "source": "plan_execution_log", # This will be part of event payload via store_knowledge
           "overall_status": final_plan_status_str,
           "user_request_preview": user_prompt_orig[:150],
           "primary_intent": plan_record.primary_intent,
           "log_timestamp_iso": plan_record.timestamp_utc # Use the record's timestamp
       }
       if plan_record.nlu_analysis_raw and plan_record.nlu_analysis_raw.get("entities"):
           for i, ent in enumerate(plan_record.nlu_analysis_raw["entities"][:3]):
               kb_meta_for_filtering[f"entity_{i+1}_type"] = ent.get("type", "UNK")
               kb_meta_for_filtering[f"entity_{i+1}_text"] = str(ent.get("text", ""))[:50]

       # Call store_knowledge with the structured content
       store_result = await self.store_knowledge(
           structured_content=plan_record,
           schema_type="PlanExecutionRecord", # Explicitly state schema type
           metadata=kb_meta_for_filtering,
           content_id=plan_record.record_id # Use the dataclass generated ID
       )

       if store_result.get("status") == "success":
           stored_kb_id = store_result.get("id") # This should be same as plan_record.record_id
           # Update the plan_record with the chroma_db_id if they are different, or ensure consistency.
           # For now, we assume content_id sets the ChromaDB ID.
           # If store_knowledge internally generates a *different* final_id for Chroma, we'd need to update plan_record.chroma_db_id
           print(f"MasterPlanner: Plan log storage successful. Stored KB ID: {stored_kb_id} (Record ID: {plan_record.record_id})")

           # The generic "kb.content.added" event is published by store_knowledge.
           # If a more specific "kb.plan_execution_log.added" event is still desired for some subscribers,
           # it could be published here additionally.
           # Example:
           # await self.publish_event(
           #     event_type="kb.plan_execution_log.added", # More specific event
           #     source_component="MasterPlanner",
           #     payload={
           #         "kb_id": stored_kb_id,
           #         "record_id": plan_record.record_id,
           #         "original_request_preview": user_prompt_orig[:150],
           #         "overall_status": final_plan_status_str,
           #         "primary_intent": plan_record.primary_intent
           #     }
           # )
           return stored_kb_id
       else:
           print(f"MasterPlanner: Plan log storage failed. Message: {store_result.get('message')}")
           return None

   async def execute_agent(self, agent: Agent, prompt: str, context: Optional[Dict] = None) -> Dict:
        print(f"Orchestrator: Executing agent {agent.name} with prompt (first 100 chars): {prompt[:100]}")
        if agent.name == "WebCrawler":
           # could be made more flexible, or we could publish that specific event here too.
           # For now, relying on the generic event from store_knowledge.
           # Example of publishing a more specific event if needed:
           # await self.publish_event(
           # event_type="kb.plan_execution_log.added",
           # source_component="MasterPlanner",
           #     payload={
           # "kb_id": stored_kb_id,
           # "original_request_preview": user_prompt_orig[:150],
           # "overall_status": final_plan_status_str,
           # "primary_intent": nlu_analysis_data.get("intent", "N/A")
           #     }
           # )
            # Add PlanLog node to Knowledge Graph & SimplifiedPlanStructure node
           if self.kg_instance and stored_kb_id: # stored_kb_id is PlanExecutionRecordDC's ID
               # Add/Update the PlanExecutionRecord node in KG
               self.kg_instance.add_node(
                   node_id=stored_kb_id,
                   node_type="PlanExecutionRecord",
                   content_preview=user_facing_summary_text[:150],
                   metadata={"status": final_plan_status_str, "intent": plan_record.primary_intent, "timestamp_utc": plan_record.timestamp_utc}
               )
               print(f"MasterPlanner: Added/Updated PlanExecutionRecord node '{stored_kb_id}' in Knowledge Graph.")

               # Create and store SimplifiedPlanStructureDC
               try:
                   parsed_plan_json = json.loads(plan_json_final_attempt) if plan_json_final_attempt else []
                   agent_sequence = [step.get("agent_name", step.get("target_agent_name", "UnknownAgent")) for step in parsed_plan_json if isinstance(step, dict)]
                   num_steps = len(parsed_plan_json)

                   key_entities_from_nlu = [entity.get("text") for entity in nlu_output_orig.get("entities", []) if entity.get("text")]
                   # Could also try to extract key nouns/concepts from user_prompt_orig if NLU entities are sparse

                   simplified_plan = SimplifiedPlanStructureDC(
                       plan_id=stored_kb_id, # Link to the main plan record ID
                       original_request_preview=user_prompt_orig[:150],
                       primary_intent=plan_record.primary_intent,
                       status=final_plan_status_str,
                       num_steps=num_steps,
                       agent_sequence=agent_sequence[:10], # Limit sequence length
                       key_abstractions_or_entities=key_entities_from_nlu[:5], # Limit entities
                       feedback_rating_if_any=None # Placeholder, feedback not directly tied here yet
                       # timestamp_utc will be set by BaseKBSchema
                   )

                   simplified_plan_node_id = f"simplan_{stored_kb_id}"
                   self.kg_instance.add_node(
                       node_id=simplified_plan_node_id,
                       node_type="SimplifiedPlan",
                       content_preview=f"Simplified plan for {plan_record.primary_intent} ({final_plan_status_str}, {num_steps} steps)",
                       metadata={"simplified_plan_json": simplified_plan.to_json_string()} # Store JSON under a key in metadata dict
                   )
                   print(f"MasterPlanner: Added/Updated SimplifiedPlan node '{simplified_plan_node_id}' in KG.")

                   # Link PlanExecutionRecord to SimplifiedPlan
                   self.kg_instance.add_edge(stored_kb_id, simplified_plan_node_id, "DESCRIBED_BY_SIMPLIFIED_PLAN", ensure_nodes=True)

                   # Link SimplifiedPlan to its primary intent topic
                   if plan_record.primary_intent:
                       topic_node_id = f"topic_{plan_record.primary_intent.lower().replace(' ', '_')}"
                       self.kg_instance.add_node(topic_node_id, "Topic", plan_record.primary_intent) # Ensure topic node exists
                       self.kg_instance.add_edge(simplified_plan_node_id, topic_node_id, "RELATED_TO_INTENT_TOPIC", ensure_nodes=True)
                   print(f"MasterPlanner: Added KG edges for SimplifiedPlan '{simplified_plan_node_id}'.")

               except Exception as e_kg_sps:
                   print(f"MasterPlanner: ERROR processing or storing SimplifiedPlanStructure in KG: {e_kg_sps}")

           return stored_kb_id
       else:
           print(f"MasterPlanner: Plan log storage failed. Message: {store_result.get('message')}")
           return None

   async def execute_agent(self, agent: Agent, prompt: str, context: Optional[Dict] = None) -> Dict:
        print(f"Orchestrator: Executing agent {agent.name} with prompt (first 100 chars): {prompt[:100]}")
        if agent.name == "WebCrawler":
            is_url = prompt.startswith("http://") or prompt.startswith("https://")
            if is_url:
                url_to_scrape = prompt; print(f"WebCrawler: Identified URL for scraping: {url_to_scrape}")
                try:
                    async with aiohttp.ClientSession() as session:
                        async with session.get(url_to_scrape, timeout=10) as response: response.raise_for_status(); html_content = await response.text()
                    soup = BeautifulSoup(html_content, 'html.parser'); paragraphs = soup.find_all('p')
                    text_content = "\n".join([p.get_text() for p in paragraphs if p.get_text()])
                    if not text_content.strip(): return {"status": "success", "agent": agent.name, "model": agent.model, "response_type": "scrape_result", "response": f"No significant text content found at {url_to_scrape}.", "summary": "", "original_url": url_to_scrape, "full_content_length": 0}
                    summary = text_content[:1000]
                    doc_summarizer_agent = next((a for a in self.agents if a.name == "DocSummarizer" and a.active), None)
                    if doc_summarizer_agent:
                        summary_prompt = f"Please summarize the following web page content obtained from {url_to_scrape}:\n\n{text_content[:15000]}"
                        # This internal _ollama_generate call could also be made an async task if summarization is slow
                        # For now, WebCrawler awaits it directly.
                        summary_result = await self._ollama_generate(doc_summarizer_agent.model, summary_prompt, context)
                        if summary_result.get("status") == "success" and summary_result.get("response"): summary = summary_result.get("response")

                    kb_metadata = {
                        "source": "web_scrape",
                        "url": url_to_scrape,
                        "scraped_timestamp_iso": datetime.datetime.utcnow().isoformat(), # Use UTC
                        "original_content_length": len(text_content),
                        "agent_used": agent.name,
                        "summarizer_used": doc_summarizer_agent.name if doc_summarizer_agent else "N/A"
                    }

                    # Create WebServiceScrapeResultDC instance
                    scrape_record = WebServiceScrapeResultDC(
                        # scrape_id and timestamp_utc are auto-generated
                        url=url_to_scrape,
                        title=soup.title.string if soup.title else None, # Basic title extraction
                        main_content_summary=summary,
                        original_content_length=len(text_content),
                        # extracted_entities could be populated by another agent/service call if needed
                        source_agent_name=agent.name
                    )
                    # Add hash of full content if we decide to implement it
                    # import hashlib
                    # scrape_record.full_content_hash = hashlib.sha256(text_content.encode('utf-8')).hexdigest()

                    # Await store_knowledge to ensure event is published if successful
                    store_kb_result = await self.store_knowledge(
                        structured_content=scrape_record,
                        schema_type="WebServiceScrapeResult", # Explicitly state schema type
                        metadata=kb_metadata, # Pass original kb_metadata as well, it contains URL, agent, etc.
                                              # store_knowledge will merge kb_schema_type into this.
                        content_id=scrape_record.scrape_id # Use dataclass generated ID
                    )

                    if store_kb_result.get("status") == "success":
                        print(f"WebCrawler: Successfully stored scraped content from {url_to_scrape} to KB (ID: {store_kb_result.get('id')}).")
                        return {"status": "success", "agent": agent.name, "model": agent.model, "response_type": "scrape_result", "response": f"Successfully scraped and summarized content from {url_to_scrape}. Summary stored in KB (ID: {store_kb_result.get('id')}).", "summary": summary, "original_url": url_to_scrape, "full_content_length": len(text_content), "kb_id": store_kb_result.get("id")}
                    else:
                        print(f"WebCrawler: Failed to store scraped content from {url_to_scrape} to KB. Error: {store_kb_result.get('message')}")
                        return {"status": "error", "agent": agent.name, "model": agent.model, "response_type": "scrape_kb_store_error", "response": f"Error storing scraped content for {url_to_scrape}: {store_kb_result.get('message')}"}

                except Exception as e:
                    return {"status": "error", "agent": agent.name, "model": agent.model, "response_type": "scrape_error", "response": f"Error scraping URL {url_to_scrape}: {str(e)}"}
            else: # WebCrawler performing a search query
                search_query = prompt
                print(f"WebCrawler: Identified search query: {search_query}")
                # DDGS is synchronous, so run in executor or make it an async task if it's too slow.
                # For now, let's assume it's fast enough or becomes an async task later.
                # If DDGS itself were async, we could await it directly.
                # To make this part non-blocking if DDGS is slow:
                # ddgs_coro = loop.run_in_executor(None, lambda: DDGS().text(keywords=search_query, region='wt-wt', max_results=5, safesearch='moderate'))
                # task_id = await self.submit_async_task(ddgs_coro, name=f"WebSearch-{agent.name}-{search_query[:20]}")
                # return {"status": "pending_async", "task_id": task_id, "message": f"Web search initiated for: {search_query}"}
                # For now, keeping it synchronous for simplicity in this step, will be a candidate for async if slow.
                try:
                    loop = asyncio.get_event_loop();
                    with DDGS() as ddgs_sync: # Renamed to avoid conflict if we make it async later
                        search_results_raw = await loop.run_in_executor(None, lambda: ddgs_sync.text(keywords=search_query, region='wt-wt', max_results=5, safesearch='moderate'))
                    formatted_results = []
                    if search_results_raw:
                        for res_raw in search_results_raw: formatted_results.append({"title": res_raw.get('title', 'N/A'), "url": res_raw.get('href', ''), "snippet": res_raw.get('body', '')})
                    return {"status": "success", "agent": agent.name, "model": agent.model, "response_type": "search_results", "response": formatted_results }
                except Exception as e:
                    return {"status": "error", "agent": agent.name, "model": agent.model, "response_type": "search_error", "response": f"Error performing web search for '{search_query}': {str(e)}"}

        elif "ollama" in agent.model:
            # This is a key change: LLM calls are now submitted as async tasks
            ollama_coro = self._ollama_generate(agent.model, prompt, context)
            task_id = await self.submit_async_task(ollama_coro, name=f"Ollama-{agent.name}-{prompt[:20]}")
            return {"status": "pending_async", "task_id": task_id, "message": f"LLM operation for agent {agent.name} started."}

        elif agent.name == "ImageForge":
            # Image generation can also be long; make it an async task.
            # generate_image_with_hf_pipeline is already async.
            image_gen_coro = self.generate_image_with_hf_pipeline(prompt)
            task_id = await self.submit_async_task(image_gen_coro, name=f"ImageForge-{prompt[:20]}")
            return {"status": "pending_async", "task_id": task_id, "message": "Image generation started."}
            # Old synchronous way: return await self.generate_image_with_hf_pipeline(prompt)

        else:
            # For other agent types not explicitly handled as async yet
            return {"status": "error", "agent": agent.name, "model": agent.model, "response": f"Execution logic for agent {agent.name} not specifically defined for async or direct execution."}

   async def classify_user_intent(self, user_prompt: str) -> Dict: # This method calls execute_agent OR self.nlu_processor
        # This method now serves as a fallback or an alternative NLU path if self.nlu_processor is not used directly in execute_master_plan
        # For direct NLUProcessor use in execute_master_plan, this method might become primarily legacy.
        # However, if it's still called, we can choose to route to NLUProcessor here too.

        # Option 1: Prioritize NLUProcessor if available and this method is called
        if self.nlu_processor:
            print("[classify_user_intent] INFO: Routing to NLUProcessor from within classify_user_intent.")
            try:
                nlu_result = self.nlu_processor.process_text(user_prompt)
                # Adapt nlu_result (NLUResult object) to the Dict structure expected by legacy callers
               # This includes adding 'parsed_command' if available.
                return {
                    "status": "success",
                    "intent": nlu_result.detected_intent.name if nlu_result.detected_intent else "unknown_intent",
                    "intent_score": nlu_result.detected_intent.confidence if nlu_result.detected_intent else 0.0,
                    "alternative_intents": [
                       {"intent": alt_name, "score": alt_score}
                       for alt_name, alt_score in (nlu_result.detected_intent.alternate_intents or [])
                    ] if nlu_result.detected_intent else [],
                    "entities": [ent.dict() for ent in nlu_result.entities],
                   "parsed_command": nlu_result.parsed_command.dict() if nlu_result.parsed_command else None, # Add parsed command
                    "implicit_goals": None, # NLUProcessor currently doesn't extract this.
                    "message": "NLU analysis via NLUProcessor (called from classify_user_intent)."
                }
            except Exception as e:
                print(f"[classify_user_intent] ERROR: NLUProcessor failed when called from classify_user_intent: {e}")
                # Fall through to legacy agent-based NLU if NLUProcessor fails here

        # Option 2: Legacy agent-based NLU (if NLUProcessor is None or failed above)
        print("[classify_user_intent] INFO: Using legacy NLUAnalysisAgent.")
        nlu_agent = next((a for a in self.agents if a.name == "NLUAnalysisAgent" and a.active), None)
        if not nlu_agent:
           return {"status": "error", "message": "NLUAnalysisAgent not found or inactive (legacy path).", "intent": None, "entities": [], "parsed_command": None, "implicit_goals": None, "alternative_intents": []}

        candidate_labels_str = ", ".join([f"'{label}'" for label in self.candidate_intent_labels])
       # Update prompt for legacy agent to also try to identify commands if possible (though less reliable than rule-based)
        nlu_prompt = (
            f"Analyze the following user prompt: '{user_prompt}'\n\n"
            f"1. Intent Classification: Classify the primary intent against candidate labels: [{candidate_labels_str}]. Provide the top intent and its confidence score (0.0-1.0).\n"
            f"2. Alternative Intents: If confidence for primary intent is below 0.85 OR the request seems ambiguous/multi-faceted, list up to 2 alternative intents with their scores.\n"
            f"3. Named Entity Recognition: Extract relevant named entities (names, locations, dates, organizations, products, filenames, URLs, specific technical terms).\n"
           f"4. Command Parsing (Experimental): If the prompt seems like a direct command (e.g., 'create file X', 'run agent Y on Z'), attempt to identify a 'command_name' (e.g., 'CREATE_FILE', 'RUN_AGENT') and 'parameters' (as a dictionary). If not a clear command, set 'command_name' to null.\n"
           f"5. Implicit Goals/Desired Outcomes: Briefly describe any underlying goals or desired outcomes the user might have, even if not explicitly stated. If none, state 'None'.\n\n"
            f"Return your analysis as a single, minified JSON object with this exact structure:\n"
            f"{{\"intent\": \"<detected_intent_label>\", \"intent_score\": <float_score_0_to_1>, "
           f"\"alternative_intents\": [{{ \"intent\": \"<alt_intent_label>\", \"score\": <float_score_0_to_1> }}], "
            f"\"entities\": [{{ \"text\": \"<entity_text>\", \"type\": \"<ENTITY_TYPE_UPPERCASE>\", \"score\": <float_score_0_to_1> }}], "
           f"\"parsed_command\": {{ \"command_name\": \"<command_or_null>\", \"parameters\": {{ \"param1\": \"value1\" }} }}, " # Added parsed_command
            f"\"implicit_goals\": \"<text_description_or_None>\"}}\n"
           f"Ensure scores are floats. If no alternatives/entities/command parameters, use empty lists/objects. If primary intent is unclear, use 'unknown_intent'."
        )
        raw_nlu_result_or_task = await self.execute_agent(nlu_agent, nlu_prompt)

        raw_nlu_result: Dict
        if raw_nlu_result_or_task.get("status") == "pending_async":
            task_id = raw_nlu_result_or_task["task_id"]
            print(f"[classify_user_intent] NLU analysis submitted as task {task_id}. Awaiting result...")
            while True:
               await asyncio.sleep(0.1)
                task_info = await self.get_async_task_info(task_id)
                if not task_info:
                   return {"status": "error", "message": f"NLU task {task_id} info not found.", "intent": None, "entities": [], "parsed_command": None}

                if task_info.status == AsyncTaskStatus.COMPLETED:
                    if not isinstance(task_info.result, dict) or "status" not in task_info.result:
                        return {"status": "error", "message": f"NLU task {task_id} completed with unexpected result format: {type(task_info.result)}.", "intent": None, "entities": [], "parsed_command": None}
                    raw_nlu_result = task_info.result
                    print(f"[classify_user_intent] NLU task {task_id} completed.")
                    break
                elif task_info.status == AsyncTaskStatus.FAILED:
                   return {"status": "error", "message": f"NLU analysis task {task_id} failed: {task_info.error}", "intent": None, "entities": [], "parsed_command": None}
                elif task_info.status == AsyncTaskStatus.CANCELLED:
                   return {"status": "error", "message": f"NLU analysis task {task_id} was cancelled.", "intent": None, "entities": [], "parsed_command": None}
        else:
            raw_nlu_result = raw_nlu_result_or_task

        if raw_nlu_result.get("status") != "success":
           return {"status": "error", "message": f"NLUAnalysisAgent call failed: {raw_nlu_result.get('response')}", "intent": None, "entities": [], "parsed_command": None, "implicit_goals": None, "alternative_intents": []}

        try:
            parsed_response = json.loads(raw_nlu_result.get("response", "{}"))
            intent = parsed_response.get("intent", "unknown_intent")
            intent_score = parsed_response.get("intent_score", 0.0)
            entities = parsed_response.get("entities", [])
            if not isinstance(entities, list): entities = []

           parsed_command_data = parsed_response.get("parsed_command")
           final_parsed_command = None
           if isinstance(parsed_command_data, dict) and parsed_command_data.get("command_name"):
               final_parsed_command = {
                   "command": parsed_command_data.get("command_name"),
                   "parameters": parsed_command_data.get("parameters", {})
               }

            alternative_intents = parsed_response.get("alternative_intents", [])
            if not isinstance(alternative_intents, list): alternative_intents = []
            implicit_goals = parsed_response.get("implicit_goals")
            if isinstance(implicit_goals, str) and implicit_goals.lower() == 'none': implicit_goals = None

           # This was for intent_scores, which isn't standard in NLUResult, so removing for now.
           # intent_scores = {intent: intent_score} if intent != "unknown_intent" else {}
           # for alt_intent_obj in alternative_intents:
           #     if isinstance(alt_intent_obj, dict) and "intent" in alt_intent_obj and "score" in alt_intent_obj:
           #         intent_scores[alt_intent_obj["intent"]] = intent_scores.get(alt_intent_obj["intent"], 0.0) + alt_intent_obj["score"]

            return {
                "status": "success",
                "intent": intent,
               "intent_score": intent_score,
                "alternative_intents": alternative_intents,
                "entities": entities,
               "parsed_command": final_parsed_command, # Include parsed command from legacy agent
                "implicit_goals": implicit_goals,
               "message": "NLU analysis via legacy agent successful."
            }
        except json.JSONDecodeError:
           return {"status": "error", "message": "NLUAnalysisAgent returned invalid JSON.", "raw_response": raw_nlu_result.get("response"), "intent": None, "entities": [], "parsed_command": None, "implicit_goals": None, "alternative_intents": []}
        except Exception as e:
           return {"status": "error", "message": f"Error processing NLU agent response: {str(e)}", "intent": None, "entities": [], "parsed_command": None, "implicit_goals": None, "alternative_intents": []}

   async def extract_document_content(self, uploaded_file_object: Any, original_filename: str) -> Dict[str, Any]:
        content_text = ""; status = "error"; message = "File type not supported or error during processing."; file_ext = Path(original_filename).suffix.lower().strip('.')
        from typing import Any
        try:
            if not hasattr(uploaded_file_object, 'getvalue'): raise ValueError("uploaded_file_object does not have getvalue() method.")
            file_bytes = uploaded_file_object.getvalue()
            if file_ext == 'txt': content_text = file_bytes.decode('utf-8', errors='replace'); status = "success"; message = "Text file processed."
            elif file_ext == 'json':
                try: parsed_json = json.loads(file_bytes.decode('utf-8', errors='replace')); content_text = json.dumps(parsed_json, indent=2); status = "success"; message = "JSON file processed."
                except json.JSONDecodeError as e: content_text = file_bytes.decode('utf-8', errors='replace'); message = f"JSON parsing error: {str(e)}. Displaying raw content."; status = "partial_success"
            elif file_ext == 'csv': content_text = file_bytes.decode('utf-8', errors='replace'); status = "success"; message = "CSV file processed as text."
            elif file_ext == 'pdf':
                try:
                    doc = fitz.open(stream=file_bytes, filetype="pdf"); text_parts = [doc.load_page(i).get_text("text") for i in range(len(doc))]; content_text = "\n".join(text_parts)
                    status = "success"; message = f"PDF processed ({len(doc)} pages).";
                    if not content_text.strip(): message += " Note: No text content found (possibly image-based PDF)."
                except Exception as e: message = f"Error processing PDF: {str(e)}"
            elif file_ext == 'docx':
                try: doc = docx.Document(io.BytesIO(file_bytes)); content_text = "\n".join([para.text for para in doc.paragraphs]); status = "success"; message = "DOCX file processed."
                except Exception as e: message = f"Error processing DOCX: {str(e)}"
            elif file_ext == 'xlsx':
                try:
                    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True); text_parts = []
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]; text_parts.append(f"--- Sheet: {sheet_name} ---")
                        for row in sheet.iter_rows(): text_parts.append(", ".join([str(cell.value) if cell.value is not None else "" for cell in row]))
                    content_text = "\n".join(text_parts); status = "success"; message = f"XLSX file processed ({len(workbook.sheetnames)} sheets)."
                except Exception as e: message = f"Error processing XLSX: {str(e)}"
            elif file_ext == 'html' or file_ext == 'htm':
                try: soup = BeautifulSoup(file_bytes, 'html.parser'); content_text = soup.get_text(separator='\n', strip=True); status = "success"; message = "HTML file processed."
                except Exception as e: message = f"Error processing HTML: {str(e)}"
            else:
                try: content_text = file_bytes.decode('utf-8', errors='replace'); status = "partial_success"; message = f"File type '{file_ext}' not specifically handled, attempted to read as text."
                except Exception: message = f"File type '{file_ext}' not supported and could not be read as raw text."; content_text = None; status = "error"
            if content_text is None: content_text = ""
        except Exception as e_outer: message = f"Outer error during file processing: {str(e_outer)}"; content_text = ""; status = "error"
        return {"status": status, "content": content_text.strip(), "file_type_processed": file_ext, "message": message}

   async def _ollama_generate(self, model_name: str, prompt: str, context: Optional[Dict] = None) -> Dict:
       print(f"Orchestrator (_ollama_generate): Calling model {model_name} with prompt: {prompt[:100]}...")
       payload = {"model": model_name, "prompt": prompt, "stream": False}
       if context: payload["context"] = context
       try:
           async with aiohttp.ClientSession() as session:
               async with session.post(self.ollama_url, json=payload) as response:
                   if response.status == 200:
                       resp_json = await response.json()
                       return {"status": "success", "response": resp_json.get("response",""), "context": resp_json.get("context")}
                   else:
                       error_text = await response.text()
                       return {"status": "error", "response": f"Ollama API error {response.status}: {error_text}"}
       except Exception as e:
           return {"status": "error", "response": f"Failed to call Ollama: {str(e)}"}

   async def generate_image_with_hf_pipeline(self, prompt:str) -> Dict[str, Any]:
        print(f"Orchestrator (ImageForge): Generating image for prompt: {prompt[:100]}...")
        time.sleep(1)
        mock_image_filename = f"generated_image_{uuid.uuid4()}.png"
        mock_image_path = self.generated_images_dir / mock_image_filename
        try:
            with open(mock_image_path, "w") as f: f.write("dummy_image_content")
            return {"status": "success", "image_path": str(mock_image_path), "response": f"Image generated and saved to {mock_image_path}"}
        except Exception as e:
            return {"status": "error", "response": f"Failed to create dummy image: {e}"}

   async def get_video_metadata(self, video_path: str) -> Dict: return {"status": "info", "message": "get_video_metadata not fully shown for brevity"}
   async def extract_video_frame(self, video_path: str, timestamp: str) -> Dict: return {"status": "info", "message": "extract_video_frame not fully shown"}
   async def convert_video_to_gif(self, video_path: str, start_time: str, end_time: str, output_filename: Optional[str]=None, scale: float=0.5, fps: int=10) -> Dict: return {"status": "info", "message": "convert_video_to_gif not fully shown"}
   async def get_audio_info(self, audio_path: str) -> Dict: return {"status": "info", "message": "get_audio_info not fully shown"}
   async def convert_audio_format(self, audio_path: str, target_format: str, output_filename_stem: Optional[str]=None) -> Dict: return {"status": "info", "message": "convert_audio_format not fully shown"}
   async def text_to_speech(self, text: str, output_filename_stem: Optional[str]=None) -> Dict: return {"status": "info", "message": "text_to_speech not fully shown"}
   async def scaffold_new_project(self, project_name: str, project_type: str) -> Dict: return {"status": "info", "message": "scaffold_new_project not fully shown"}
   async def explain_code_snippet(self, code_snippet: str, language: Optional[str] = "python") -> Dict: return {"status": "info", "message": "explain_code_snippet not fully shown"}
   async def generate_code_module(self, requirements: str, language: Optional[str] = "python", filename: Optional[str] = None) -> Dict: return {"status": "info", "message": "generate_code_module not fully shown"}
   async def get_system_info(self, component: str) -> Dict: return {"status": "info", "message": "get_system_info not fully shown"}
   async def parallel_execution(self, prompt: str, selected_agents_names: List[str], context: Optional[Dict]=None) -> List[Dict]: return [{"status":"info", "message":"parallel_execution not fully shown"}]

   async def _service_docprocessor_process_and_store(self, params: Dict, service_definition: AgentServiceDefinition) -> Dict:
        """
        Service handler for DocProcessor's 'process_and_store_document' service.
        Summarizes document text, stores it as GenericDocumentDC in KB, and returns KB ID.
        """
        log_prefix = f"[{self.__class__.__name__}._service_docprocessor_process_and_store]"
        document_text = params.get("document_text")
        source_identifier = params.get("source_identifier")
        summarization_instructions = params.get("summarization_instructions") # Optional

        if not document_text or not isinstance(document_text, str) or not document_text.strip():
            return {"status": "error", "data": None, "message": "Missing or empty 'document_text' parameter.", "error_code": "MISSING_PARAMETER_DOCUMENT_TEXT"}
        if not source_identifier or not isinstance(source_identifier, str) or not source_identifier.strip():
            return {"status": "error", "data": None, "message": "Missing or empty 'source_identifier' parameter.", "error_code": "MISSING_PARAMETER_SOURCE_ID"}

        summarizer_agent = next((a for a in self.agents if a.name == "DocSummarizer" and a.active), None)
        if not summarizer_agent:
            return {"status": "error", "data": None, "message": "DocSummarizer agent not available for summarization.", "error_code": "AGENT_UNAVAILABLE_DOCSUMMARIZER"}

        # Construct prompt for DocSummarizer
        max_input_length = 15000 # Consistent with _service_docsummarizer_summarize_text
        summary_prompt = f"Please provide a concise summary of the following text from source '{source_identifier}'.\n"
        if summarization_instructions:
            summary_prompt += f"Follow these instructions for summarization: {summarization_instructions}\n"
        summary_prompt += f"\nTEXT TO SUMMARIZE:\n```\n{document_text[:max_input_length]}\n```\n\nCONCISE SUMMARY:"

        print(f"{log_prefix} Calling DocSummarizer. Text length: {len(document_text)} (truncated to {max_input_length} for prompt), Source: '{source_identifier}'.")

        # Use the existing direct service call logic if DocSummarizer.summarize_text has a direct handler
        # This avoids creating a new async task if the underlying service can handle it.
        # However, DocSummarizer.summarize_text itself calls execute_agent which creates an async task.
        # So, we will call execute_agent directly here to manage the async task lifecycle properly for this service.

        summarization_llm_call_or_task = await self.execute_agent(summarizer_agent, summary_prompt)

        summary_text: Optional[str] = None

        if summarization_llm_call_or_task.get("status") == "pending_async":
            task_id = summarization_llm_call_or_task["task_id"]
            print(f"{log_prefix} Summarization task {task_id} submitted for document '{source_identifier}'. Awaiting result...")

            # This service handler needs to await the result of its own async sub-task.
            # The _handle_agent_service_call method expects this handler to return the final result,
            # or propagate 'pending_async' if the handler itself wants to be non-blocking (which is not the case here, we need the summary).
            while True:
                await asyncio.sleep(0.2) # Poll interval
                task_info = await self.get_async_task_info(task_id)
                if not task_info:
                    msg = f"Summarization LLM task {task_id} info not found for document '{source_identifier}'."
                    print(f"{log_prefix} ERROR: {msg}")
                    return {"status": "error", "data": None, "message": msg, "error_code": "ASYNC_TASK_NOT_FOUND_SUMMARY"}

                if task_info.status == AsyncTaskStatus.COMPLETED:
                    if task_info.result and task_info.result.get("status") == "success":
                        summary_text = task_info.result.get("response")
                        print(f"{log_prefix} Summarization LLM task {task_id} for '{source_identifier}' completed. Summary length: {len(summary_text) if summary_text else 0}")
                    else:
                        msg = f"Summarization LLM task {task_id} for '{source_identifier}' completed but failed or no response: {task_info.result}"
                        print(f"{log_prefix} ERROR: {msg}")
                        return {"status": "error", "data": None, "message": msg, "error_code": "LLM_SUMMARY_TASK_FAILED_OR_EMPTY"}
                    break
                elif task_info.status == AsyncTaskStatus.FAILED:
                    msg = f"Summarization LLM task {task_id} for '{source_identifier}' failed: {task_info.error}"
                    print(f"{log_prefix} ERROR: {msg}")
                    return {"status": "error", "data": None, "message": msg, "error_code": "LLM_SUMMARY_TASK_FAILED"}
                elif task_info.status == AsyncTaskStatus.CANCELLED:
                    msg = f"Summarization LLM task {task_id} for '{source_identifier}' was cancelled."
                    print(f"{log_prefix} WARN: {msg}") # Treat as error for this service's purpose
                    return {"status": "error", "data": None, "message": msg, "error_code": "LLM_SUMMARY_TASK_CANCELLED"}

        elif summarization_llm_call_or_task.get("status") == "success": # Synchronous success from execute_agent (less likely for LLM)
            summary_text = summarization_llm_call_or_task.get("response")
            print(f"{log_prefix} Summarization for '{source_identifier}' completed synchronously (unexpected for LLM). Summary length: {len(summary_text) if summary_text else 0}")
        else: # Synchronous error from execute_agent
            error_msg = summarization_llm_call_or_task.get("response", f"Unknown error during summarization agent call for '{source_identifier}'.")
            print(f"{log_prefix} ERROR: Summarization agent call for '{source_identifier}' failed (sync path): {error_msg}")
            return {"status": "error", "data": None, "message": error_msg, "error_code": "LLM_SUMMARY_CALL_FAILED_SYNC"}

        if not summary_text or not summary_text.strip():
            msg = f"Summarization for document '{source_identifier}' resulted in an empty summary."
            print(f"{log_prefix} ERROR: {msg}")
            return {"status": "error", "data": None, "message": msg, "error_code": "EMPTY_SUMMARY_GENERATED"}

        # Create GenericDocumentDC instance
        doc_processor_agent_name = service_definition.name.split('.')[0] # Infer from service name "DocProcessor.process..."

        generic_doc_record = GenericDocumentDC(
            source_identifier=source_identifier,
            original_content=document_text,
            summary_content=summary_text,
            processing_notes=f"Summarized with instructions: '{summarization_instructions}'" if summarization_instructions else "Standard summarization.",
            source_agent_name=doc_processor_agent_name # The agent providing this service
        )

        print(f"{log_prefix} Storing GenericDocument for '{source_identifier}' in KB. Record ID: {generic_doc_record.record_id}")
        store_result = await self.store_knowledge(
            structured_content=generic_doc_record,
            schema_type="GenericDocument", # Explicitly state schema type
            metadata={ # Basic metadata for top-level filtering
                "source": "doc_processor_service",
                "doc_source_id": source_identifier[:100], # Truncate for safety as metadata value
                "agent_responsible": doc_processor_agent_name
            },
            content_id=generic_doc_record.record_id
        )

        if store_result.get("status") == "success":
            kb_id = store_result.get("id")
            success_msg = f"Document '{source_identifier}' processed, summarized, and stored in KB. KB ID: {kb_id}"
            print(f"{log_prefix} SUCCESS: {success_msg}")
            return {
                "status": "success",
                "data": {"kb_id": kb_id, "summary_preview": summary_text[:150] + "..."},
                "message": success_msg
            }
        else:
            error_msg = f"Failed to store processed document '{source_identifier}' in KB. Error: {store_result.get('message')}"
            print(f"{log_prefix} ERROR: {error_msg}")
            return {"status": "error", "data": None, "message": error_msg, "error_code": "KB_STORE_FAILED_GENERIC_DOC"}

orchestrator = TerminusOrchestrator()
